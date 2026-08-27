# -*- coding: utf-8 -*-
"""URT 回测任务持久化（urt_backtest_tasks）。"""

from __future__ import annotations

import csv
import io
import logging
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

from sqlalchemy import desc
from sqlalchemy.orm import Session

from backend_api.database import SessionLocal
from backend_api.models import URTBacktestTask

logger = logging.getLogger(__name__)

# 明细导出列顺序与中文表头（打开 CSV/Excel 时显示）
_URT_DETAIL_FIELDS: List[str] = [
    "code",
    "name",
    "signal_date",
    "score",
    "entry_date",
    "entry_price",
    "max_high",
    "max_gain_pct",
    "hit_target",
    "hit_date",
    "hit_target_lower",
    "hit_target_upper",
    "hit_date_upper",
    "hit_in_band",
    "observation_end_date",
    "horizon_days",
    "exit_date",
    "exit_price",
    "exit_reason",
    "pnl_pct",
    "bars_held",
    "structure_fallback",
    "fallback_reason",
    "structure_source",
    "stop_basis",
    "target_basis",
    "nearest_support",
    "nearest_resistance",
    "stop_price",
    "target_price",
    "kde_ok",
    "partial_exit",
    "partial_frac",
    "partial_exit_price",
    "partial_exit_date",
    "exit_reason_combo",
    "f_above_ma20",
    "f_yang",
    "f_yang_quality",
    "f_volume",
    "f_yang_medium",
    "f_ma_bull",
    "f_turnover",
    "f_volume_ratio",
    "f_structure_position",
    "f_overheat_penalty",
    "volume_multiple",
    "yang_count_5",
    "structure_rr",
    "dist_to_support_pct",
    "proximity_reason",
    "turnover_rate",
    "turnover_relative",
    "ma_bull_depth",
    "overheat_intensity",
    "ret_from_low_n",
    "horizon_pnl_pct",
    "horizon_exit_price",
]

_URT_DETAIL_HEADER_ZH: Dict[str, str] = {
    "code": "股票代码",
    "name": "股票名称",
    "signal_date": "信号日期",
    "score": "得分",
    "entry_date": "入场日期",
    "entry_price": "入场价",
    "max_high": "观察期最高价",
    "max_gain_pct": "观察期最大涨幅(%)",
    "hit_target": "是否命中目标",
    "hit_date": "命中日期",
    "hit_target_lower": "是否触及下限",
    "hit_target_upper": "是否触及上限",
    "hit_date_upper": "触及上限日期",
    "hit_in_band": "涨幅落在区间内",
    "observation_end_date": "观察期结束日",
    "horizon_days": "观察期天数",
    "exit_date": "出场日期",
    "exit_price": "出场价(期末收盘)",
    "exit_reason": "出场原因",
    "pnl_pct": "期末盈亏比例(%)",
    "bars_held": "持有天数",
    "structure_fallback": "结构回退",
    "fallback_reason": "回退原因",
    "structure_source": "结构来源",
    "stop_basis": "止损依据",
    "target_basis": "止盈依据",
    "nearest_support": "最近支撑",
    "nearest_resistance": "最近阻力",
    "stop_price": "止损价",
    "target_price": "止盈价",
    "kde_ok": "KDE有效",
    "partial_exit": "分批出场",
    "partial_frac": "分批比例",
    "partial_exit_price": "分批出场价",
    "partial_exit_date": "分批出场日",
    "exit_reason_combo": "组合出场原因",
    "f_above_ma20": "MA20趋势分",
    "f_yang": "连阳分",
    "f_yang_quality": "阳线质量分",
    "f_volume": "量能分",
    "f_yang_medium": "中期阳线分",
    "f_ma_bull": "均线多头分",
    "f_turnover": "换手分",
    "f_volume_ratio": "量比分",
    "f_structure_position": "结构位分",
    "f_overheat_penalty": "过热扣分",
    "volume_multiple": "量能倍数",
    "yang_count_5": "5日连阳",
    "structure_rr": "结构盈亏比",
    "dist_to_support_pct": "距支撑(%)",
    "proximity_reason": "贴近支撑原因",
    "turnover_rate": "换手率(%)",
    "turnover_relative": "换手相对中位",
    "ma_bull_depth": "均线多头深度",
    "overheat_intensity": "过热强度",
    "ret_from_low_n": "近低点涨幅",
    "horizon_pnl_pct": "满观察期盈亏(%)",
    "horizon_exit_price": "满观察期收盘",
}

_URT_EXIT_REASON_ZH: Dict[str, str] = {
    "target_hit": "触及目标",
    "horizon_end": "到期平仓",
    "rule_exit": "规则离场",
    "stop_loss": "止损",
    "price_stop": "价格止损",
    "time_stop": "时间止损",
    "trailing_drawdown": "回撤止盈",
    "trailing_take_profit": "回撤止盈",
    "structure_stop": "结构止损",
    "structure_target": "阻力止盈",
    "pct_target": "百分比止盈",
    "breakeven_stop": "保本止损",
    "fallback_trail": "移动止盈",
}

# xlsx 列宽（中文表头）；未知列由内容自适应
_URT_BACKTEST_XLSX_COL_WIDTH: Dict[str, float] = {
    "股票代码": 12.0,
    "股票名称": 14.0,
    "信号日期": 13.0,
    "得分": 8.0,
    "入场日期": 13.0,
    "入场价": 10.0,
    "观察期最高价": 14.0,
    "观察期最大涨幅(%)": 18.0,
    "是否命中目标": 14.0,
    "命中日期": 13.0,
    "是否触及下限": 14.0,
    "是否触及上限": 14.0,
    "触及上限日期": 14.0,
    "涨幅落在区间内": 16.0,
    "观察期结束日": 14.0,
    "观察期天数": 12.0,
    "出场日期": 13.0,
    "出场价(期末收盘)": 16.0,
    "出场原因": 12.0,
    "期末盈亏比例(%)": 16.0,
    "持有天数": 10.0,
    "结构回退": 10.0,
    "回退原因": 14.0,
    "结构来源": 12.0,
    "止损依据": 12.0,
    "止盈依据": 12.0,
    "最近支撑": 10.0,
    "最近阻力": 10.0,
    "止损价": 10.0,
    "止盈价": 10.0,
    "KDE有效": 10.0,
    "分批出场": 10.0,
    "分批比例": 10.0,
    "分批出场价": 12.0,
    "分批出场日": 13.0,
    "组合出场原因": 14.0,
    "MA20趋势分": 12.0,
    "量能倍数": 10.0,
    "5日连阳": 10.0,
    "结构盈亏比": 12.0,
    "距支撑(%)": 12.0,
    "贴近支撑原因": 16.0,
    "换手率(%)": 12.0,
    "换手相对中位": 14.0,
    "均线多头深度": 14.0,
    "过热强度": 12.0,
    "近低点涨幅": 12.0,
    "满观察期盈亏(%)": 16.0,
    "满观察期收盘": 14.0,
}

_URT_XLSX_DATE_HEADERS = frozenset(
    {
        "信号日期",
        "入场日期",
        "命中日期",
        "触及上限日期",
        "观察期结束日",
        "出场日期",
        "分批出场日",
    }
)


def _excel_text_display_width(value: Any) -> float:
    """估算 Excel 列宽（CJK 约 2 单位，ASCII 约 1）。"""
    s = str(value or "").replace(_EXCEL_TEXT_PREFIX, "")
    w = 0.0
    for ch in s:
        w += 2.2 if ord(ch) > 127 else 1.0
    return w


def _autofit_urt_xlsx_columns(ws: Any, headers: List[str], data_rows: List[List[str]]) -> None:
    from openpyxl.utils import get_column_letter

    sample = data_rows[:300]
    for col_idx, header in enumerate(headers, start=1):
        base = _URT_BACKTEST_XLSX_COL_WIDTH.get(header, 12.0)
        content_w = _excel_text_display_width(header)
        for row in sample:
            if col_idx - 1 < len(row):
                content_w = max(content_w, _excel_text_display_width(row[col_idx - 1]))
        width = max(base, content_w + 2.0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 8.0), 42.0)


def _apply_urt_xlsx_cell_formats(ws: Any, headers: List[str]) -> None:
    """日期列按文本显示，避免 #######；股票代码列保持文本。"""
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    code_col = col_map.get("股票代码")
    date_cols = [col_map[h] for h in _URT_XLSX_DATE_HEADERS if h in col_map]
    if ws.max_row < 2:
        return
    for row in range(2, ws.max_row + 1):
        if code_col:
            cell = ws.cell(row=row, column=code_col)
            cell.number_format = "@"
            if cell.value is not None:
                cell.value = str(cell.value)
        for dc in date_cols:
            cell = ws.cell(row=row, column=dc)
            cell.number_format = "@"
            if cell.value is not None and cell.value != "":
                cell.value = str(cell.value)[:10]


def _session() -> Session:
    return SessionLocal()


def normalize_task_id(task_id: Optional[str]) -> str:
    return str(task_id or "").strip()


def clamp_progress(progress: Any) -> int:
    try:
        v = int(round(float(progress)))
    except (TypeError, ValueError):
        return 0
    return max(0, min(100, v))


def _dt_iso(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat() + "Z" if dt else None


def _row_to_dict(row: URTBacktestTask) -> Dict[str, Any]:
    return {
        "task_id": row.task_id,
        "name": row.name,
        "config": row.config if isinstance(row.config, dict) else {},
        "status": row.status,
        "progress": clamp_progress(row.progress),
        "message": row.message or "",
        "logs": row.logs if isinstance(row.logs, list) else [],
        "created_at": _dt_iso(row.created_at),
        "started_at": _dt_iso(row.started_at),
        "completed_at": _dt_iso(row.completed_at),
        "summary": row.summary,
        "details_path": row.details_path,
        "error": row.error,
        "has_details_csv": bool(row.details_csv_bytes),
    }


def ensure_table(db: Optional[Session] = None) -> None:
    owns = db is None
    if owns:
        db = _session()
    try:
        URTBacktestTask.__table__.create(bind=db.get_bind(), checkfirst=True)
    finally:
        if owns:
            db.close()


def create_task(config: Dict[str, Any], name: Optional[str] = None) -> str:
    ensure_table()
    task_id = str(uuid.uuid4())
    now = datetime.utcnow()
    db = _session()
    try:
        t = URTBacktestTask(
            task_id=task_id,
            name=name or config.get("task_name") or f"URT回测_{task_id[:8]}",
            config=config,
            status="pending",
            progress=0,
            message="",
            logs=[],
            summary=None,
            error=None,
            details_path=None,
            details_csv_bytes=None,
            created_at=now,
            started_at=None,
            completed_at=None,
        )
        db.add(t)
        db.commit()
        return task_id
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def get_task(task_id: str) -> Optional[Dict[str, Any]]:
    tid = normalize_task_id(task_id)
    if not tid:
        return None
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        return _row_to_dict(row) if row else None
    finally:
        db.close()


def list_tasks(limit: int = 50, status: Optional[str] = None) -> List[Dict[str, Any]]:
    db = _session()
    try:
        q = db.query(URTBacktestTask).order_by(desc(URTBacktestTask.created_at))
        if status:
            q = q.filter(URTBacktestTask.status == status)
        return [_row_to_dict(r) for r in q.limit(int(limit)).all()]
    finally:
        db.close()


def update_task_progress(task_id: str, progress: int, message: str = "", log_line: Optional[str] = None) -> None:
    tid = normalize_task_id(task_id)
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row:
            return
        if row.status == "pending":
            row.status = "running"
            row.started_at = datetime.utcnow()
        row.progress = clamp_progress(progress)
        row.message = message or row.message
        if log_line:
            logs = list(row.logs or [])
            logs.append({"ts": datetime.utcnow().isoformat() + "Z", "message": log_line})
            row.logs = logs[-200:]
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def _urt_detail_field_order(rows: List[Dict[str, Any]]) -> List[str]:
    """稳定列顺序：优先预定义字段，其余追加在后。"""
    seen = set()
    ordered: List[str] = []
    for k in _URT_DETAIL_FIELDS:
        ordered.append(k)
        seen.add(k)
    if rows:
        for k in rows[0].keys():
            if k not in seen:
                ordered.append(k)
                seen.add(k)
    return ordered


# 与 report_service URT 日报一致：零宽字符前缀使 Excel 按文本显示并保留前导零
_EXCEL_TEXT_PREFIX = "\u2060"


def _normalize_export_stock_code(value: Any) -> str:
    """导出用股票代码：A 股补零为 6 位；5 位港股码不 zfill(6)；加零宽前缀防 Excel 丢前导零。"""
    s = str(value or "").replace(_EXCEL_TEXT_PREFIX, "").strip()
    if not s:
        return ""
    if s.isdigit():
        # 5 位数字是港股码，禁止抬成 6 位 A 股（00981→000981）
        if len(s) != 5 and 0 < len(s) <= 6:
            s = s.zfill(6)
    return _EXCEL_TEXT_PREFIX + s


def _format_urt_detail_cell(key: str, value: Any) -> Any:
    if key in ("hit_target", "hit_target_lower", "hit_target_upper", "hit_in_band"):
        if value is True or value == 1 or str(value).lower() in ("true", "1", "yes"):
            return "是"
        if value is False or value == 0 or str(value).lower() in ("false", "0", "no"):
            return "否"
        return value
    if key == "exit_reason":
        s = str(value or "").strip()
        return _URT_EXIT_REASON_ZH.get(s, s or "")
    if key == "code" and value is not None:
        return _normalize_export_stock_code(value)
    return value


def _build_urt_details_csv_bytes(details_rows: List[Dict[str, Any]]) -> bytes:
    fields = _urt_detail_field_order(details_rows)
    headers_zh = [_URT_DETAIL_HEADER_ZH.get(k, k) for k in fields]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=headers_zh, extrasaction="ignore", quoting=csv.QUOTE_MINIMAL)
    w.writeheader()
    for r in details_rows:
        row_zh = {
            _URT_DETAIL_HEADER_ZH.get(k, k): _format_urt_detail_cell(k, r.get(k))
            for k in fields
        }
        w.writerow(row_zh)
    return buf.getvalue().encode("utf-8-sig")


def complete_task(
    task_id: str,
    summary: Dict[str, Any],
    details_rows: Optional[List[Dict[str, Any]]] = None,
) -> None:
    tid = normalize_task_id(task_id)
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row:
            return
        row.status = "completed"
        row.progress = 100
        row.message = "完成"
        row.summary = summary
        row.completed_at = datetime.utcnow()
        row.error = None
        if details_rows:
            row.details_csv_bytes = _build_urt_details_csv_bytes(details_rows)
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def patch_task_config(task_id: str, extra: Dict[str, Any]) -> None:
    """合并写入任务 config JSON（对照任务 ID 等）。"""
    tid = normalize_task_id(task_id)
    if not tid or not extra:
        return
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row:
            return
        cfg = dict(row.config) if isinstance(row.config, dict) else {}
        cfg.update(extra)
        row.config = cfg
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def patch_task_summary(task_id: str, extra: Dict[str, Any]) -> None:
    """合并写入任务 summary JSON。"""
    tid = normalize_task_id(task_id)
    if not tid or not extra:
        return
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row:
            return
        summary = dict(row.summary) if isinstance(row.summary, dict) else {}
        summary.update(extra)
        row.summary = summary
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def fail_task(task_id: str, error: str) -> None:
    tid = normalize_task_id(task_id)
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row:
            return
        row.status = "failed"
        row.error = error
        row.message = error
        row.completed_at = datetime.utcnow()
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def cancel_task(task_id: str) -> bool:
    tid = normalize_task_id(task_id)
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row or row.status in ("completed", "failed", "cancelled"):
            return False
        row.status = "cancelled"
        row.message = "已取消"
        row.completed_at = datetime.utcnow()
        db.commit()
        return True
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def delete_task(task_id: str) -> bool:
    tid = normalize_task_id(task_id)
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row:
            return False
        db.delete(row)
        db.commit()
        return True
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def _rewrite_csv_stock_codes_excel_text(raw: bytes) -> bytes:
    """下载/转 xlsx 前规范化「股票代码」列（兼容历史未补零/无零宽前缀的 CSV）。"""
    if not raw:
        return raw
    text = raw.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return raw
    headers = rows[0]
    try:
        code_idx = headers.index("股票代码")
    except ValueError:
        return raw
    out = io.StringIO()
    w = csv.writer(out, quoting=csv.QUOTE_MINIMAL)
    w.writerow(headers)
    for row in rows[1:]:
        row = list(row)
        if len(row) > code_idx:
            row[code_idx] = _normalize_export_stock_code(row[code_idx])
        w.writerow(row)
    return out.getvalue().encode("utf-8-sig")


def get_details_csv(task_id: str) -> Optional[bytes]:
    tid = normalize_task_id(task_id)
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row or not row.details_csv_bytes:
            return None
        return _rewrite_csv_stock_codes_excel_text(bytes(row.details_csv_bytes))
    finally:
        db.close()


def _csv_bytes_to_xlsx(raw: bytes) -> bytes:
    """将已存 CSV 明细转为 xlsx（保留中文表头；股票代码按文本写入）。"""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    text = raw.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return b""
    wb = Workbook()
    ws = wb.active
    ws.title = "URT回测明细"
    headers = rows[0]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    code_col = None
    for i, h in enumerate(headers, start=1):
        if h == "股票代码":
            code_col = i
            break
    for row in rows[1:]:
        if code_col and len(row) >= code_col:
            # 与 URT 日报一致：规范化 + 零宽前缀，并保证写入为文本
            idx = code_col - 1
            row = list(row)
            row[idx] = _normalize_export_stock_code(row[idx])
        ws.append(row)
    _apply_urt_xlsx_cell_formats(ws, headers)
    _autofit_urt_xlsx_columns(ws, headers, rows[1:])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_details_xlsx(task_id: str) -> Optional[bytes]:
    raw = get_details_csv(task_id)
    if not raw:
        return None
    return _csv_bytes_to_xlsx(raw)


def count_completed_reports() -> int:
    db = _session()
    try:
        return db.query(URTBacktestTask).filter(URTBacktestTask.status == "completed").count()
    finally:
        db.close()


def count_running_tasks() -> int:
    db = _session()
    try:
        return (
            db.query(URTBacktestTask)
            .filter(URTBacktestTask.status.in_(("pending", "running")))
            .count()
        )
    finally:
        db.close()


def list_reports(limit: int = 50, offset: int = 0) -> List[Dict[str, Any]]:
    """已完成任务投影为报告列表。"""
    db = _session()
    try:
        rows = (
            db.query(URTBacktestTask)
            .filter(URTBacktestTask.status == "completed")
            .order_by(desc(URTBacktestTask.completed_at), desc(URTBacktestTask.created_at))
            .offset(int(offset))
            .limit(int(limit))
            .all()
        )
        out = []
        for r in rows:
            d = _row_to_dict(r)
            out.append(
                {
                    "report_id": d["task_id"],
                    "task_id": d["task_id"],
                    "name": d["name"],
                    "created_at": d.get("completed_at") or d.get("created_at"),
                    "summary": d.get("summary"),
                    "details_path": d.get("details_path"),
                    "has_details_csv": d.get("has_details_csv"),
                    "config": d.get("config"),
                }
            )
        return out
    finally:
        db.close()


def get_report(report_id: str) -> Optional[Dict[str, Any]]:
    row = get_task(report_id)
    if not row or row.get("status") != "completed":
        return None
    return {
        "report_id": row["task_id"],
        "task_id": row["task_id"],
        "name": row["name"],
        "created_at": row.get("completed_at") or row.get("created_at"),
        "summary": row.get("summary"),
        "details_path": row.get("details_path"),
        "has_details_csv": row.get("has_details_csv"),
        "config": row.get("config"),
    }


def get_task_logs(task_id: str) -> List[Dict[str, Any]]:
    row = get_task(task_id)
    if not row:
        return []
    logs = row.get("logs") or []
    # 统一为 {text, ts} 便于前端
    out = []
    for item in logs:
        if isinstance(item, dict):
            text = item.get("message") or item.get("text") or str(item)
            out.append({"text": text, "ts": item.get("ts")})
        else:
            out.append({"text": str(item)})
    return out


def reset_task_for_rerun(task_id: str) -> bool:
    tid = normalize_task_id(task_id)
    db = _session()
    try:
        row = db.query(URTBacktestTask).filter(URTBacktestTask.task_id == tid).first()
        if not row:
            return False
        row.status = "pending"
        row.progress = 0
        row.message = "重新排队"
        row.error = None
        row.summary = None
        row.details_csv_bytes = None
        row.started_at = None
        row.completed_at = None
        logs = list(row.logs or [])
        logs.append({"ts": datetime.utcnow().isoformat() + "Z", "message": "任务重新执行"})
        row.logs = logs[-200:]
        db.commit()
        return True
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def batch_delete_tasks(task_ids: List[str]) -> int:
    n = 0
    for tid in task_ids or []:
        if delete_task(tid):
            n += 1
    return n
