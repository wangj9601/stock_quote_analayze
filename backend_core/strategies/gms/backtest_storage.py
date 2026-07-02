"""
GMS 回测任务与报告持久化（PostgreSQL gms_backtest_tasks）
"""

import csv
import io
import logging
import math
import threading
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import desc, nulls_last
from sqlalchemy.orm import Session

from backend_api.database import SessionLocal
from backend_api.models import GMSBacktestTask

logger = logging.getLogger(__name__)

_INDEX_LOCK = threading.Lock()


def _session() -> Session:
    return SessionLocal()


def normalize_gms_task_id(task_id: Optional[str]) -> str:
    """
    规范化 URL/参数中的 task_id：去首尾空白，并将各类 Unicode 连字符统一为 ASCII '-'。
    """
    if task_id is None:
        return ""
    s = str(task_id).strip()
    if not s:
        return ""
    for ch in ("\u2010", "\u2011", "\u2012", "\u2013", "\u2014", "\u2015", "\u2212", "\uff0d"):
        s = s.replace(ch, "-")
    return s


def clamp_gms_progress(progress: Any) -> int:
    """任务进度限制在 0–100"""

    try:
        v = int(round(float(progress)))
    except (TypeError, ValueError):
        return 0
    return max(0, min(100, v))


def _dt_to_iso_z(dt: Optional[datetime]) -> Optional[str]:
    if dt is None:
        return None
    return dt.isoformat() + "Z"


def _row_to_dict(row: GMSBacktestTask) -> Dict[str, Any]:
    logs = row.logs if isinstance(row.logs, list) else []
    cfg = row.config if isinstance(row.config, dict) else {}
    return {
        "task_id": row.task_id,
        "name": row.name,
        "config": cfg,
        "status": row.status,
        "progress": clamp_gms_progress(row.progress),
        "message": row.message or "",
        "logs": logs,
        "created_at": _dt_to_iso_z(row.created_at),
        "started_at": _dt_to_iso_z(row.started_at),
        "completed_at": _dt_to_iso_z(row.completed_at),
        "summary": row.summary,
        "details_path": row.details_path,
        "error": row.error,
    }


def create_task(config: Dict[str, Any], name: Optional[str] = None) -> str:
    """创建任务记录，返回 task_id。"""
    task_id = str(uuid.uuid4())
    now = datetime.utcnow()
    db = _session()
    try:
        t = GMSBacktestTask(
            task_id=task_id,
            name=name or config.get("task_name") or f"GMS回测_{task_id[:8]}",
            config=config,
            status="pending",
            progress=0,
            message="",
            logs=[],
            summary=None,
            error=None,
            details_path=None,
            details_csv_bytes=None,
            details_xlsx_bytes=None,
            created_at=now,
            started_at=None,
            completed_at=None,
        )
        db.add(t)
        db.commit()
        return task_id
    except Exception as e:
        db.rollback()
        logger.exception("创建 GMS 任务失败: %s", e)
        raise
    finally:
        db.close()


def get_task(task_id: str) -> Optional[Dict[str, Any]]:
    """获取任务详情。"""
    tid = normalize_gms_task_id(task_id)
    if not tid:
        return None
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
        if not row:
            logger.debug("任务不存在: %r", task_id)
            return None
        return _row_to_dict(row)
    finally:
        db.close()


def update_task_progress(task_id: str, progress: int, message: str = "", log_line: Optional[str] = None) -> bool:
    """更新任务进度与可选日志。"""
    tid = normalize_gms_task_id(task_id)
    if not tid:
        return False
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
        if not row:
            return False
        row.progress = clamp_gms_progress(progress)
        row.message = message
        if log_line is not None:
            logs = list(row.logs) if isinstance(row.logs, list) else []
            logs.append({"ts": datetime.utcnow().isoformat() + "Z", "text": log_line})
            row.logs = logs
        if row.status == "pending" and progress > 0:
            row.status = "running"
            if row.started_at is None:
                row.started_at = datetime.utcnow()
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        logger.warning("更新任务进度失败 %s: %s", tid, e)
        return False
    finally:
        db.close()


def append_task_log(task_id: str, log_line: str) -> bool:
    """仅追加一条日志。"""
    tid = normalize_gms_task_id(task_id)
    if not tid:
        return False
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
        if not row:
            return False
        logs = list(row.logs) if isinstance(row.logs, list) else []
        logs.append({"ts": datetime.utcnow().isoformat() + "Z", "text": log_line})
        row.logs = logs
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        logger.warning("追加任务日志失败 %s: %s", tid, e)
        return False
    finally:
        db.close()


def complete_task(task_id: str, summary: Dict[str, Any], details_path: Optional[str] = None) -> bool:
    """标记任务完成并写入汇总与明细路径（逻辑文件名）。"""
    tid = normalize_gms_task_id(task_id)
    if not tid:
        return False
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
        if not row:
            return False
        row.status = "completed"
        row.progress = 100
        row.completed_at = datetime.utcnow()
        row.summary = summary
        row.details_path = details_path
        row.error = None
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        logger.warning("完成任务写入失败 %s: %s", tid, e)
        return False
    finally:
        db.close()


def fail_task(task_id: str, error: str) -> bool:
    """标记任务失败。"""
    tid = normalize_gms_task_id(task_id)
    if not tid:
        return False
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
        if not row:
            return False
        row.status = "failed"
        row.error = error
        row.completed_at = datetime.utcnow()
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        logger.warning("失败任务写入失败 %s: %s", tid, e)
        return False
    finally:
        db.close()


def cancel_task(task_id: str) -> bool:
    """标记任务已取消。"""
    tid = normalize_gms_task_id(task_id)
    if not tid:
        return False
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
        if not row:
            return False
        if row.status in ("completed", "failed"):
            return False
        row.status = "cancelled"
        row.completed_at = datetime.utcnow()
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        logger.warning("取消任务写入失败 %s: %s", tid, e)
        return False
    finally:
        db.close()


def delete_task(task_id: str) -> bool:
    """删除任务记录。"""
    tid = normalize_gms_task_id(task_id)
    if not tid:
        return False
    with _INDEX_LOCK:
        db = _session()
        try:
            row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
            if row:
                db.delete(row)
                db.commit()
            return True
        except Exception as e:
            db.rollback()
            logger.warning("删除任务失败 %s: %s", tid, e)
            return False
        finally:
            db.close()


def list_tasks(status: Optional[str] = None, limit: int = 50, offset: int = 0) -> List[Dict[str, Any]]:
    """任务列表，按创建时间倒序。"""
    db = _session()
    try:
        q = db.query(GMSBacktestTask).order_by(GMSBacktestTask.created_at.desc())
        if status:
            q = q.filter(GMSBacktestTask.status == status)
        rows = q.offset(offset).limit(limit).all()
        return [_row_to_dict(r) for r in rows]
    finally:
        db.close()


def get_task_logs(task_id: str) -> List[Dict[str, Any]]:
    """返回任务日志列表。"""
    task = get_task(task_id)
    if not task:
        return []
    return task.get("logs") or []


def list_reports(limit: int = 50, offset: int = 0) -> List[Dict[str, Any]]:
    """报告列表（已完成任务），按完成时间倒序。"""
    db = _session()
    try:
        rows = (
            db.query(GMSBacktestTask)
            .filter(GMSBacktestTask.status == "completed")
            .order_by(nulls_last(desc(GMSBacktestTask.completed_at)))
            .offset(offset)
            .limit(limit)
            .all()
        )
        out = []
        for row in rows:
            tid = row.task_id
            out.append(
                {
                    "report_id": tid,
                    "task_id": tid,
                    "name": row.name,
                    "created_at": _dt_to_iso_z(row.completed_at),
                    "summary": row.summary,
                    "details_path": row.details_path,
                }
            )
        return out
    finally:
        db.close()


def delete_report(report_id: str) -> bool:
    """删除已完成任务的报告（删除整条任务记录及库内明细）。"""
    rid = normalize_gms_task_id(report_id)
    if not rid:
        return False
    with _INDEX_LOCK:
        db = _session()
        try:
            row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == rid).first()
            if not row or row.status != "completed":
                return False
            db.delete(row)
            db.commit()
            return True
        except Exception as e:
            db.rollback()
            logger.warning("删除报告失败 %s: %s", rid, e)
            return False
        finally:
            db.close()


def get_report(report_id: str) -> Optional[Dict[str, Any]]:
    """报告详情。"""
    rid = normalize_gms_task_id(report_id)
    if not rid:
        return None
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == rid).first()
        if not row or row.status != "completed":
            return None
        tid = row.task_id
        return {
            "report_id": tid,
            "task_id": tid,
            "name": row.name,
            "created_at": _dt_to_iso_z(row.completed_at),
            "summary": row.summary,
            "details_path": row.details_path,
        }
    finally:
        db.close()


def get_details_path(report_id: str) -> Optional[str]:
    """逻辑明细文件名（优先 xlsx）。"""
    rid = normalize_gms_task_id(report_id)
    if not rid:
        return None
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == rid).first()
        if not row:
            return None
        if row.details_path:
            return row.details_path
        if row.details_xlsx_bytes:
            return f"{rid}.xlsx"
        if row.details_csv_bytes:
            return f"{rid}.csv"
        return None
    finally:
        db.close()


def get_detail_path_by_ext(report_id: str, ext: str) -> Optional[str]:
    """兼容旧接口：明细已入库，不再返回磁盘路径。"""
    return None


def get_report_file_bytes(
    report_id: str, variant: Optional[str] = None
) -> Optional[Tuple[bytes, str, str]]:
    """
    返回报告明细字节与下载文件名、Content-Type。
    variant: None 优先 xlsx；csv / xlsx 强制格式。
    """
    rid = normalize_gms_task_id(report_id)
    if not rid:
        return None
    v = (variant or "").strip().lower()
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == rid).first()
        if not row or row.status != "completed":
            return None
        base_name = (row.name or f"gms_backtest_{rid[:8]}").strip()
        safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in base_name) or f"gms_backtest_{rid[:8]}"

        def pick() -> Optional[Tuple[bytes, str, str]]:
            if v == "csv":
                if not row.details_csv_bytes:
                    return None
                return (
                    row.details_csv_bytes,
                    f"{safe_name}.csv",
                    "text/csv; charset=utf-8",
                )
            if v in ("xlsx", "excel"):
                if not row.details_xlsx_bytes:
                    return None
                data = row.details_xlsx_bytes
                if row.summary and isinstance(row.summary, dict):
                    try:
                        data = _inject_summary_into_xlsx(data, row.summary)
                    except Exception as e:
                        logger.warning("注入统计摘要 sheet 失败: %s", e)
                return (
                    data,
                    f"{safe_name}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            if row.details_xlsx_bytes:
                data = row.details_xlsx_bytes
                if row.summary and isinstance(row.summary, dict):
                    try:
                        data = _inject_summary_into_xlsx(data, row.summary)
                    except Exception as e:
                        logger.warning("注入统计摘要 sheet 失败: %s", e)
                return (
                    data,
                    f"{safe_name}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            if row.details_csv_bytes:
                return (row.details_csv_bytes, f"{safe_name}.csv", "text/csv; charset=utf-8")
            return None

        return pick()
    finally:
        db.close()


def count_completed_reports() -> int:
    db = _session()
    try:
        return db.query(GMSBacktestTask).filter(GMSBacktestTask.status == "completed").count()
    finally:
        db.close()


def count_running_tasks() -> int:
    db = _session()
    try:
        return (
            db.query(GMSBacktestTask)
            .filter(GMSBacktestTask.status.in_(("pending", "running")))
            .count()
        )
    finally:
        db.close()


def normalize_gms_stock_code(code: Any, market: Any = None) -> str:
    """
    将股票代码规范为字符串：港股纯数字不足 5 位前补零，A 股不足 6 位前补零。
    """
    if code is None:
        return ""
    if isinstance(code, float):
        if math.isnan(code):
            return ""
        if code.is_integer():
            code = int(code)
    s = str(code).strip()
    if not s:
        return ""
    if not s.isdigit():
        return s
    mt = str(market or "").upper()
    if "HK" in mt:
        if len(s) < 5:
            return s.zfill(5)
    elif "CN" in mt:
        if len(s) < 6:
            return s.zfill(6)
    else:
        if len(s) < 5:
            return s.zfill(5)
    return s


def format_code_for_csv_cell(code: Any, market: Any = None) -> str:
    norm = normalize_gms_stock_code(code, market)
    if not norm:
        return ""
    return "\t" + norm


_GMS_BACKTEST_DETAIL_CSV_HEADER_ZH: Dict[str, str] = {
    "code": "股票代码",
    "date": "信号日期",
    "market": "市场",
    "buy_type": "买点类型",
    "score_total": "信号总分",
    "entry_open": "买入价",
    "entry_close": "买入价",
    "max_high_20d": "观察期内最高价",
    "max_gain_20d": "观察期内最大涨幅",
    "hit": "是否命中目标",
    "entry_date": "入场日期",
    "entry_exec_price": "入场成交价",
    "exit_date": "出场日期",
    "exit_price": "出场价格",
    "exit_exec_price": "出场成交价",
    "exit_reason": "出场原因",
    "bars_held": "持有K线数",
    "pnl_pct": "单笔收益率",
    "position_fraction": "单笔仓位比例",
    "portfolio_pnl_pct": "按仓位计收益率",
    "stop_loss_pct": "止损阈值",
    "commission_bps": "手续费bps",
    "slippage_bps": "滑点bps",
    "atr_period": "ATR周期",
    "init_stop_atr_k": "初始止损ATR倍数",
    "trail_stop_mode": "移动止损模式",
    "trail_atr_k": "跟踪ATR倍数",
    "trail_pct": "回撤止损比例",
    "breakeven_trigger_r": "保本触发R",
    "profit_lock_trigger_r": "锁盈触发R",
    "profit_lock_r": "锁盈保留R",
    "partial_take_profit_r": "分批止盈触发R",
    "partial_take_profit_applied": "是否触发分批止盈",
    "partial_take_ratio": "分批止盈比例",
    "r_multiple": "R倍数",
    "initial_risk_pct": "初始风险比例",
    "initial_stop_price": "初始止损价",
    "max_favorable_excursion_pct": "最大有利波动",
    "max_adverse_excursion_pct": "最大不利波动",
}


def _gms_detail_csv_fieldnames_zh(keys: List[str]) -> List[str]:
    return [_GMS_BACKTEST_DETAIL_CSV_HEADER_ZH.get(k, k) for k in keys]


def _gms_detail_row_to_zh_row(row: Dict[str, Any], keys: List[str]) -> Dict[str, Any]:
    zh_keys = _gms_detail_csv_fieldnames_zh(keys)
    return {zh_keys[i]: row.get(keys[i]) for i in range(len(keys))}


_GMS_DETAIL_DEFAULT_FIELDS: List[str] = [
    "code",
    "date",
    "market",
    "buy_type",
    "score_total",
    "entry_open",
    "max_high_20d",
    "max_gain_20d",
    "hit",
]

_GMS_BACKTEST_XLSX_COL_WIDTH: Dict[str, float] = {
    "股票代码": 11.0,
    "信号日期": 12.0,
    "市场": 8.0,
    "买点类型": 10.0,
    "信号总分": 10.0,
    "买入价": 12.0,
    "观察期内最高价": 16.0,
    "观察期内最大涨幅": 18.0,
    "是否命中目标": 12.0,
    "入场日期": 12.0,
    "入场成交价": 12.0,
    "出场日期": 12.0,
    "出场价格": 12.0,
    "出场成交价": 12.0,
    "出场原因": 12.0,
    "持有K线数": 12.0,
    "单笔收益率": 12.0,
    "单笔仓位比例": 12.0,
    "按仓位计收益率": 14.0,
    "止损阈值": 12.0,
    "手续费bps": 12.0,
    "滑点bps": 12.0,
    "ATR周期": 10.0,
    "初始止损ATR倍数": 14.0,
    "移动止损模式": 14.0,
    "跟踪ATR倍数": 12.0,
    "回撤止损比例": 12.0,
    "保本触发R": 10.0,
    "锁盈触发R": 10.0,
    "锁盈保留R": 10.0,
    "分批止盈触发R": 14.0,
    "是否触发分批止盈": 14.0,
    "分批止盈比例": 12.0,
    "R倍数": 10.0,
    "初始风险比例": 12.0,
    "初始止损价": 12.0,
    "最大有利波动": 12.0,
    "最大不利波动": 12.0,
}


def _build_gms_detail_rows(
    details: List[Dict[str, Any]],
    *,
    code_csv_format: bool = False,
) -> Tuple[List[str], List[str], List[Dict[str, Any]]]:
    if not details:
        keys = list(_GMS_DETAIL_DEFAULT_FIELDS)
        fieldnames_zh = _gms_detail_csv_fieldnames_zh(keys)
        return keys, fieldnames_zh, []
    keys = list(details[0].keys())
    fieldnames_zh = _gms_detail_csv_fieldnames_zh(keys)
    rows_out = []
    for raw in details:
        r = dict(raw)
        if "code" in r:
            if code_csv_format:
                r["code"] = format_code_for_csv_cell(r.get("code"), r.get("market"))
            else:
                r["code"] = normalize_gms_stock_code(r.get("code"), r.get("market"))
        rows_out.append(r)
    rows_zh = [_gms_detail_row_to_zh_row(r, keys) for r in rows_out]
    return keys, fieldnames_zh, rows_zh


def _sort_gms_details_for_export(details: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not details:
        return []
    return sorted(
        details,
        key=lambda d: (
            0 if (d.get("market") or "") == "CN" else 1,
            str(d.get("code") or ""),
            str(d.get("date") or ""),
        ),
    )


def _gms_rows_zh_insert_blank_between_codes(
    fieldnames_zh: List[str], rows_zh: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    if not rows_zh:
        return []
    out: List[Dict[str, Any]] = []
    prev: Optional[str] = None
    for row in rows_zh:
        c = row.get("股票代码")
        c_cmp = str(c).lstrip("\t") if c is not None else ""
        if prev is not None and c_cmp != prev:
            out.append({fn: "" for fn in fieldnames_zh})
        prev = c_cmp
        out.append(row)
    return out


def _write_gms_xlsx_sheet(ws: Any, fieldnames_zh: List[str], rows_zh: List[Dict[str, Any]]) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if not fieldnames_zh:
        return
    ws.append(list(fieldnames_zh))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hit_header = "是否命中目标"
    prev_code: Optional[str] = None
    for row_dict in rows_zh:
        c = row_dict.get("股票代码")
        c_cmp = str(c).lstrip("\t") if c is not None else ""
        if prev_code is not None and c_cmp != prev_code:
            ws.append([""] * len(fieldnames_zh))
        prev_code = c_cmp
        row_vals = []
        for h in fieldnames_zh:
            v = row_dict.get(h)
            if h == hit_header:
                if v is True:
                    v = "是"
                elif v is False:
                    v = "否"
            row_vals.append(v)
        ws.append(row_vals)

    for col_idx, header in enumerate(fieldnames_zh, start=1):
        w = _GMS_BACKTEST_XLSX_COL_WIDTH.get(header, 14.0)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    code_col_idx = None
    for i, h in enumerate(fieldnames_zh, start=1):
        if h == "股票代码":
            code_col_idx = i
            break
    if code_col_idx and ws.max_row >= 2:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=code_col_idx).number_format = "@"

    gain_header = "观察期内最大涨幅"
    gain_col_idx = None
    for i, h in enumerate(fieldnames_zh, start=1):
        if h == gain_header:
            gain_col_idx = i
            break
    if gain_col_idx and ws.max_row >= 2:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=gain_col_idx)
            val = cell.value
            if val is not None and val != "" and isinstance(val, (int, float)):
                cell.number_format = "0.00%"
    for header in ("单笔收益率", "按仓位计收益率", "止损阈值", "单笔仓位比例"):
        pct_col_idx = None
        for i, h in enumerate(fieldnames_zh, start=1):
            if h == header:
                pct_col_idx = i
                break
        if pct_col_idx and ws.max_row >= 2:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=pct_col_idx)
                val = cell.value
                if val is not None and val != "" and isinstance(val, (int, float)):
                    cell.number_format = "0.00%"


def _build_csv_bytes(details: List[Dict[str, Any]]) -> bytes:
    sorted_details = _sort_gms_details_for_export(details)
    _, fieldnames_zh, rows_zh = _build_gms_detail_rows(sorted_details, code_csv_format=True)
    hit_hdr = "是否命中目标"
    for row in rows_zh:
        if hit_hdr in row and isinstance(row[hit_hdr], bool):
            row[hit_hdr] = "是" if row[hit_hdr] else "否"
    rows_out = _gms_rows_zh_insert_blank_between_codes(fieldnames_zh, rows_zh)
    buffer = io.StringIO()
    w = csv.DictWriter(buffer, fieldnames=fieldnames_zh, quoting=csv.QUOTE_ALL)
    w.writeheader()
    w.writerows(rows_out)
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def _write_summary_xlsx_sheet(ws: Any, summary: Dict[str, Any]) -> None:
    from openpyxl.styles import Font

    ws.title = "统计摘要"
    ws.append(["GMS 回测统计摘要"])
    ws["A1"].font = Font(bold=True, size=14)
    version = summary.get("summary_schema_version")
    if version:
        ws.append(["摘要版本", version])
    for key, label in (
        ("hit_rate", "目标命中率"),
        ("win_rate", "胜率"),
        ("profit_factor", "盈亏比"),
        ("max_drawdown", "最大回撤"),
        ("approx_annual_return_simple", "近似年化收益"),
        ("trade_count", "交易笔数"),
        ("signal_count", "信号数"),
    ):
        if summary.get(key) is not None:
            ws.append([label, summary[key]])

    hist = summary.get("holding_days_histogram") or {}
    if hist:
        ws.append([])
        ws.append(["持有天数分布"])
        ws.append(["区间", "笔数"])
        for bucket, cnt in hist.items():
            ws.append([bucket, cnt])

    monthly = summary.get("monthly_returns") or []
    if monthly:
        ws.append([])
        ws.append(["分月收益"])
        ws.append(["月份", "收益率%", "笔数"])
        for m in monthly:
            ws.append([m.get("month"), m.get("return_pct"), m.get("trade_count")])

    by_sig = summary.get("by_signal_type") or {}
    if by_sig:
        ws.append([])
        ws.append(["信号类型对比"])
        ws.append(["类型", "胜率", "平均R", "笔数"])
        for sig, stats in by_sig.items():
            if isinstance(stats, dict):
                ws.append([sig, stats.get("win_rate"), stats.get("avg_r"), stats.get("trade_count")])


def _build_xlsx_bytes(details: List[Dict[str, Any]], summary: Optional[Dict[str, Any]] = None) -> bytes:
    from openpyxl import Workbook

    sorted_details = _sort_gms_details_for_export(details)
    _, fieldnames_zh, rows_zh = _build_gms_detail_rows(sorted_details, code_csv_format=False)
    cn_rows = [r for r in rows_zh if r.get("市场") == "CN"]
    etf_rows = [r for r in rows_zh if r.get("市场") == "ETF"]
    hk_rows = [r for r in rows_zh if r.get("市场") == "HK"]

    wb = Workbook()
    wb.remove(wb.active)
    sheet_idx = 0
    if summary:
        ws_sum = wb.create_sheet("统计摘要", sheet_idx)
        _write_summary_xlsx_sheet(ws_sum, summary)
        sheet_idx += 1
    ws_cn = wb.create_sheet("A股", sheet_idx)
    _write_gms_xlsx_sheet(ws_cn, fieldnames_zh, cn_rows)
    ws_etf = wb.create_sheet("ETF", sheet_idx + 1)
    _write_gms_xlsx_sheet(ws_etf, fieldnames_zh, etf_rows)
    ws_hk = wb.create_sheet("港股", sheet_idx + 2)
    _write_gms_xlsx_sheet(ws_hk, fieldnames_zh, hk_rows)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _inject_summary_into_xlsx(xlsx_bytes: bytes, summary: Dict[str, Any]) -> bytes:
    """为已生成的 XLSX 注入统计摘要 sheet（下载时兼容旧任务）。"""
    from openpyxl import load_workbook

    bio_in = io.BytesIO(xlsx_bytes)
    wb = load_workbook(bio_in)
    if "统计摘要" in wb.sheetnames:
        return xlsx_bytes
    ws_sum = wb.create_sheet("统计摘要", 0)
    _write_summary_xlsx_sheet(ws_sum, summary)
    bio_out = io.BytesIO()
    wb.save(bio_out)
    return bio_out.getvalue()


def save_details_csv(task_id: str, details: List[Dict[str, Any]]) -> str:
    """生成 CSV 字节并写入任务行。"""
    tid = normalize_gms_task_id(task_id) or str(task_id).strip()
    fname = f"{tid}.csv"
    data = _build_csv_bytes(details)
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
        if row:
            row.details_csv_bytes = data
            db.commit()
        return fname
    except Exception as e:
        db.rollback()
        logger.warning("保存 CSV 明细失败 %s: %s", tid, e)
        raise
    finally:
        db.close()


def save_details_xlsx(task_id: str, details: List[Dict[str, Any]], summary: Optional[Dict[str, Any]] = None) -> str:
    """生成 Excel 字节并写入任务行。"""
    tid = normalize_gms_task_id(task_id) or str(task_id).strip()
    fname = f"{tid}.xlsx"
    data = _build_xlsx_bytes(details, summary=summary)
    db = _session()
    try:
        row = db.query(GMSBacktestTask).filter(GMSBacktestTask.task_id == tid).first()
        if row:
            row.details_xlsx_bytes = data
            db.commit()
        return fname
    except Exception as e:
        db.rollback()
        logger.warning("保存 XLSX 明细失败 %s: %s", tid, e)
        raise
    finally:
        db.close()
