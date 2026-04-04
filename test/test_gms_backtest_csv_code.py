"""GMS 回测明细 CSV 中 code 列规范化（港股补零、Excel 文本）"""

import os
import sys
import tempfile

project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from backend_core.strategies.gms.backtest_storage import (
    normalize_gms_stock_code,
    normalize_gms_task_id,
    format_code_for_csv_cell,
    save_details_csv,
    save_details_xlsx,
)


def test_normalize_task_id_unicode_hyphens():
    good = "cfaa27e6-d72a-4751-9e5a-a76224e4ea21"
    bad = good.replace("-", "\u2011")  # 非断行连字符，常见于复制粘贴
    assert normalize_gms_task_id(bad) == good
    assert normalize_gms_task_id("  " + good + "  ") == good


def test_normalize_hk_short_code():
    assert normalize_gms_stock_code(981, "HK") == "00981"
    assert normalize_gms_stock_code("981", "HK") == "00981"
    assert normalize_gms_stock_code(981.0, "HK") == "00981"
    assert normalize_gms_stock_code("00981", "HK") == "00981"


def test_normalize_cn_short_code():
    assert normalize_gms_stock_code(1, "CN") == "000001"
    assert normalize_gms_stock_code("000001", "CN") == "000001"


def test_format_code_for_csv_excel_tab():
    assert format_code_for_csv_cell(981, "HK") == "\t00981"
    assert format_code_for_csv_cell("000001", "CN").startswith("\t")


def test_save_details_csv_code_column(monkeypatch):
    from backend_core.strategies.gms import backtest_storage as bs

    d = tempfile.mkdtemp(prefix="gms_csv_")
    try:
        monkeypatch.setattr(bs, "_DETAILS_DIR", d)
        tid = "test-task-csv-1"
        details = [
            {
                "code": 981,
                "date": "2024-01-02",
                "market": "HK",
                "buy_type": "左侧",
                "score_total": 80.0,
                "entry_open": 10.0,
                "max_high_20d": 11.0,
                "max_gain_20d": 0.1,
                "hit": True,
            }
        ]
        fname = save_details_csv(tid, details)
        path = os.path.join(d, fname)
        with open(path, "r", encoding="utf-8-sig") as f:
            lines = f.readlines()
        assert len(lines) >= 2
        assert "\t00981" in lines[1]
        assert "是" in lines[1]  # 是否命中目标与 Excel 一致为中文
    finally:
        try:
            os.remove(os.path.join(d, f"{tid}.csv"))
        except OSError:
            pass
        try:
            os.rmdir(d)
        except OSError:
            pass


def test_save_details_xlsx_chinese_headers_and_code_text(monkeypatch):
    from backend_core.strategies.gms import backtest_storage as bs
    from openpyxl import load_workbook

    d = tempfile.mkdtemp(prefix="gms_xlsx_")
    try:
        monkeypatch.setattr(bs, "_DETAILS_DIR", d)
        tid = "test-task-xlsx-1"
        details = [
            {
                "code": 981,
                "date": "2024-01-02",
                "market": "HK",
                "buy_type": "左侧",
                "score_total": 80.0,
                "entry_open": 10.0,
                "max_high_20d": 11.0,
                "max_gain_20d": 0.1,
                "hit": True,
            }
        ]
        fname = save_details_xlsx(tid, details)
        assert fname == f"{tid}.xlsx"
        path = os.path.join(d, fname)
        wb = load_workbook(path, data_only=True)
        assert "A股" in wb.sheetnames and "港股" in wb.sheetnames
        ws_cn = wb["A股"]
        assert ws_cn["A1"].value == "股票代码"
        assert ws_cn.max_row == 1
        ws_hk = wb["港股"]
        assert ws_hk["A2"].value == "00981"
        assert ws_hk.cell(row=2, column=ws_hk.max_column).value == "是"
        dim = ws_hk.column_dimensions["A"]
        assert dim.width is not None and dim.width >= 10
        wb.close()
    finally:
        try:
            os.remove(os.path.join(d, f"{tid}.xlsx"))
        except OSError:
            pass
        try:
            os.rmdir(d)
        except OSError:
            pass
