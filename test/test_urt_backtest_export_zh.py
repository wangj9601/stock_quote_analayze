"""URT 回测明细导出中文列名与股票代码 Excel 文本处理。"""

import sys
import os
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend_core.strategies.urt.backtest_storage import (
    _EXCEL_TEXT_PREFIX,
    _build_urt_details_csv_bytes,
    _csv_bytes_to_xlsx,
    _normalize_export_stock_code,
    _rewrite_csv_stock_codes_excel_text,
)


def test_urt_details_csv_chinese_headers():
    rows = [
        {
            "code": "000676",
            "name": "智度股份",
            "signal_date": "2026-07-17",
            "score": 86.0,
            "entry_date": "2026-07-18",
            "entry_price": 6.8,
            "exit_date": "2026-07-20",
            "exit_price": 7.5,
            "exit_reason": "target_hit",
            "hit_target": True,
            "hit_date": "2026-07-20",
            "pnl_pct": 10.2,
            "bars_held": 2,
        }
    ]
    text = _build_urt_details_csv_bytes(rows).decode("utf-8-sig")
    header = text.splitlines()[0]
    assert "股票代码" in header
    assert "股票名称" in header
    assert "信号日期" in header
    assert "是否命中目标" in header
    assert "出场原因" in header
    assert "量能分" in header
    assert "满观察期盈亏(%)" in header
    assert "code" not in header
    body = text.splitlines()[1]
    assert "是" in body
    assert "触及目标" in body
    assert "000676" in body


def test_normalize_export_stock_code_a_share_zfill_and_excel_prefix():
    assert _normalize_export_stock_code("700") == _EXCEL_TEXT_PREFIX + "000700"
    assert _normalize_export_stock_code("2036") == _EXCEL_TEXT_PREFIX + "002036"
    assert _normalize_export_stock_code("000011") == _EXCEL_TEXT_PREFIX + "000011"
    assert _normalize_export_stock_code(_EXCEL_TEXT_PREFIX + "11") == _EXCEL_TEXT_PREFIX + "000011"


def test_normalize_export_stock_code_keeps_hk_5_digit():
    """港股 5 位码不得 zfill(6) 抬成 A 股。"""
    assert _normalize_export_stock_code("00981") == _EXCEL_TEXT_PREFIX + "00981"
    assert _normalize_export_stock_code("981") == _EXCEL_TEXT_PREFIX + "000981"  # 3 位按 A 股补零


def test_urt_details_csv_stock_code_excel_text():
    rows = [
        {"code": "700", "name": "模塑科技", "signal_date": "2026-07-17", "score": 80.0, "hit_target": True},
        {"code": "000676", "name": "智度股份", "signal_date": "2026-07-17", "score": 86.0, "hit_target": False},
    ]
    text = _build_urt_details_csv_bytes(rows).decode("utf-8-sig")
    lines = text.splitlines()
    assert _EXCEL_TEXT_PREFIX + "000700" in lines[1]
    assert _EXCEL_TEXT_PREFIX + "000676" in lines[2]
    # 裸数字码不应以无前缀形式单独出现在单元格起始（Excel 会当数值）
    assert not lines[1].startswith("700,")
    assert not lines[1].startswith("000700,")


def test_rewrite_legacy_csv_stock_codes_for_excel():
    """历史 CSV 无零宽前缀/未补零时，下载侧仍可规范化。"""
    legacy = "股票代码,股票名称\n700,模塑科技\n981,中芯国际\n".encode("utf-8-sig")
    fixed = _rewrite_csv_stock_codes_excel_text(legacy).decode("utf-8-sig")
    lines = fixed.splitlines()
    assert lines[1].startswith(_EXCEL_TEXT_PREFIX + "000700")
    # 981 为 3 位 → A 股补零；真正 5 位港股码另测
    assert lines[2].startswith(_EXCEL_TEXT_PREFIX + "000981")
    hk = "股票代码,股票名称\n00981,中芯国际-H\n".encode("utf-8-sig")
    hk_fixed = _rewrite_csv_stock_codes_excel_text(hk).decode("utf-8-sig")
    assert _EXCEL_TEXT_PREFIX + "00981" in hk_fixed.splitlines()[1]
    assert "000981" not in hk_fixed


def test_urt_details_xlsx_from_csv():
    rows = [{"code": "000676", "name": "智度股份", "signal_date": "2026-07-17", "score": 86.0, "hit_target": True}]
    raw = _build_urt_details_csv_bytes(rows)
    xlsx = _csv_bytes_to_xlsx(raw)
    assert xlsx[:2] == b"PK"
    assert len(xlsx) > 100


def test_urt_details_xlsx_keeps_leading_zeros():
    from openpyxl import load_workbook

    rows = [
        {"code": "700", "name": "模塑科技", "signal_date": "2026-07-17", "score": 80.0, "hit_target": True},
        {"code": "2036", "name": "联创电子", "signal_date": "2026-07-17", "score": 70.0, "hit_target": False},
    ]
    xlsx = _csv_bytes_to_xlsx(_build_urt_details_csv_bytes(rows))
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    code_col = headers.index("股票代码") + 1
    c1 = ws.cell(row=2, column=code_col)
    c2 = ws.cell(row=3, column=code_col)
    assert c1.number_format == "@"
    assert str(c1.value).replace(_EXCEL_TEXT_PREFIX, "") == "000700"
    assert str(c2.value).replace(_EXCEL_TEXT_PREFIX, "") == "002036"
    assert str(c1.value).startswith(_EXCEL_TEXT_PREFIX)
