"""
测试 ReportService - 报告生成服务
"""

import os
import pytest
import pandas as pd
from datetime import datetime
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from backend_api.models import Base, User, Watchlist, HistoricalQuotes, HistoricalQuotesHK
from backend_api.services.report_service import ReportService, ReportInfo, ReportResult


# 测试数据库配置
TEST_DATABASE_URL = "postgresql+psycopg2://postgres:qidianspacetime@localhost:5446/stock_analysis"


@pytest.fixture(scope="module")
def test_engine():
    """创建测试数据库引擎"""
    engine = create_engine(TEST_DATABASE_URL, echo=False)
    return engine


@pytest.fixture(scope="function")
def test_db(test_engine):
    """创建测试数据库会话"""
    Session = sessionmaker(bind=test_engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture(scope="function")
def test_user(test_db):
    """创建测试用户"""
    # 检查用户是否已存在
    existing_user = test_db.query(User).filter(User.username == "test_report_user").first()
    if existing_user:
        return existing_user
    
    user = User(
        username="test_report_user",
        email="test_report@example.com",
        password_hash="test_hash",
        role="user",
        status="active"
    )
    test_db.add(user)
    test_db.commit()
    test_db.refresh(user)
    return user


@pytest.fixture(scope="function")
def test_watchlist(test_db, test_user):
    """创建测试自选股"""
    # 清理已存在的测试自选股
    test_db.query(Watchlist).filter(Watchlist.user_id == test_user.id).delete()
    test_db.commit()
    
    # 添加测试自选股（A股和港股）
    watchlist_items = [
        Watchlist(
            user_id=test_user.id,
            stock_code="000001",
            stock_name="平安银行",
            group_name="default"
        ),
        Watchlist(
            user_id=test_user.id,
            stock_code="600000",
            stock_name="浦发银行",
            group_name="default"
        ),
        Watchlist(
            user_id=test_user.id,
            stock_code="00700",
            stock_name="腾讯控股",
            group_name="default"
        )
    ]
    
    for item in watchlist_items:
        test_db.add(item)
    
    test_db.commit()
    return watchlist_items


@pytest.fixture(scope="function")
def report_service(test_db):
    """创建 ReportService 实例"""
    return ReportService(test_db, report_dir="test_reports")


def test_get_user_watchlist_empty(report_service, test_db):
    """测试获取空自选股列表"""
    # 检查用户是否已存在
    existing_user = test_db.query(User).filter(User.username == "test_empty_user").first()
    if existing_user:
        # 清理已存在用户的自选股
        test_db.query(Watchlist).filter(Watchlist.user_id == existing_user.id).delete()
        test_db.commit()
        user = existing_user
    else:
        # 创建一个没有自选股的用户
        user = User(
            username="test_empty_user",
            email="test_empty@example.com",
            password_hash="test_hash"
        )
        test_db.add(user)
        test_db.commit()
        test_db.refresh(user)
    
    watchlist = report_service.get_user_watchlist(user.id)
    assert watchlist == []
    
    # 清理
    test_db.delete(user)
    test_db.commit()


def test_get_user_watchlist_with_data(report_service, test_user, test_watchlist):
    """测试获取有数据的自选股列表"""
    watchlist = report_service.get_user_watchlist(test_user.id)
    
    assert len(watchlist) == 3
    assert watchlist[0]['stock_code'] == "000001"
    assert watchlist[0]['stock_name'] == "平安银行"
    assert watchlist[0]['market'] == "CN"
    
    assert watchlist[2]['stock_code'] == "00700"
    assert watchlist[2]['stock_name'] == "腾讯控股"
    assert watchlist[2]['market'] == "HK"


def test_get_user_watchlist_with_filter(report_service, test_user, test_watchlist):
    """测试获取指定股票的自选股列表"""
    watchlist = report_service.get_user_watchlist(test_user.id, stock_codes=["000001"])
    
    assert len(watchlist) == 1
    assert watchlist[0]['stock_code'] == "000001"


def test_determine_market(report_service):
    """测试市场类型判断"""
    assert report_service._determine_market("000001") == "CN"
    assert report_service._determine_market("600000") == "CN"
    assert report_service._determine_market("00700") == "HK"
    assert report_service._determine_market("01810") == "HK"


def test_get_stock_history_data(report_service, test_db):
    """测试获取股票历史数据"""
    # 测试A股历史数据
    history_data = report_service.get_stock_history_data("000001", "CN", days=5)
    
    if history_data:
        assert len(history_data) > 0
        assert 'trade_date' in history_data[0]
        assert 'open_price' in history_data[0]
        assert 'close_price' in history_data[0]
        print(f"✅ 获取A股历史数据成功，共 {len(history_data)} 条")
    else:
        print("⚠️ A股历史数据为空（可能数据库中没有数据）")
    
    # 测试港股历史数据
    history_data_hk = report_service.get_stock_history_data("00700", "HK", days=5)
    
    if history_data_hk:
        assert len(history_data_hk) > 0
        assert 'trade_date' in history_data_hk[0]
        print(f"✅ 获取港股历史数据成功，共 {len(history_data_hk)} 条")
    else:
        print("⚠️ 港股历史数据为空（可能数据库中没有数据）")


def test_get_stock_summary_data(report_service, test_db):
    """测试获取股票汇总数据"""
    # 测试A股汇总数据
    summary = report_service.get_stock_summary_data("000001", "CN")
    
    if summary:
        assert 'stock_name' in summary
        assert 'current_price' in summary
        assert 'market' in summary
        assert summary['market'] == "CN"
        print(f"✅ 获取A股汇总数据成功: {summary['stock_name']}")
    else:
        print("⚠️ A股汇总数据为空（可能数据库中没有数据）")
    
    # 测试港股汇总数据
    summary_hk = report_service.get_stock_summary_data("00700", "HK")
    
    if summary_hk:
        assert 'stock_name' in summary_hk
        assert summary_hk['market'] == "HK"
        print(f"✅ 获取港股汇总数据成功: {summary_hk['stock_name']}")
    else:
        print("⚠️ 港股汇总数据为空（可能数据库中没有数据）")


def test_generate_summary_report_empty_watchlist(report_service, test_db):
    """测试生成空自选股的汇总报告"""
    # 检查用户是否已存在
    existing_user = test_db.query(User).filter(User.username == "test_empty_report_user").first()
    if existing_user:
        # 清理已存在用户的自选股
        test_db.query(Watchlist).filter(Watchlist.user_id == existing_user.id).delete()
        test_db.commit()
        user = existing_user
    else:
        # 创建一个没有自选股的用户
        user = User(
            username="test_empty_report_user",
            email="test_empty_report@example.com",
            password_hash="test_hash"
        )
        test_db.add(user)
        test_db.commit()
        test_db.refresh(user)
    
    result = report_service.generate_user_report(user.id, 'summary')
    
    assert result.success is True
    assert result.file_path is None
    assert result.report_info.stock_count == 0
    assert result.report_info.has_data is False
    assert result.error_message == "用户没有自选股"
    
    # 清理
    test_db.delete(user)
    test_db.commit()


def test_generate_summary_report_with_data(report_service, test_user, test_watchlist):
    """测试生成有数据的汇总报告"""
    result = report_service.generate_user_report(test_user.id, 'summary')
    
    assert result.success is True
    assert result.report_info.stock_count == 3
    assert result.report_info.report_type == 'summary'
    
    if result.file_path:
        assert os.path.exists(result.file_path)
        assert result.file_path.endswith('.csv')
        
        # 读取CSV文件验证内容
        df = pd.read_csv(result.file_path, encoding='utf-8-sig')
        assert len(df) == 3
        assert '股票代码' in df.columns
        assert '股票名称' in df.columns
        assert '市场' in df.columns
        
        print(f"✅ 生成汇总报告成功: {result.file_path}")
        print(f"   股票数量: {result.report_info.stock_count}")
        print(f"   文件大小: {result.report_info.file_size} 字节")
        print(f"   数据缺失股票: {result.report_info.missing_data_stocks}")
        
        # 清理测试文件
        os.remove(result.file_path)
    else:
        print("⚠️ 未生成报告文件（可能数据库中没有数据）")


def test_generate_detailed_report_with_data(report_service, test_user, test_watchlist):
    """测试生成有数据的详细报告"""
    result = report_service.generate_user_report(test_user.id, 'detailed')
    
    assert result.success is True
    assert result.report_info.stock_count == 3
    assert result.report_info.report_type == 'detailed'
    
    if result.file_path:
        assert os.path.exists(result.file_path)
        assert result.file_path.endswith('.xlsx')
        
        # 读取Excel文件验证内容
        df_summary = pd.read_excel(result.file_path, sheet_name='股票汇总')
        assert len(df_summary) == 3
        
        df_history = pd.read_excel(result.file_path, sheet_name='历史数据')
        assert len(df_history) > 0
        
        print(f"✅ 生成详细报告成功: {result.file_path}")
        print(f"   股票数量: {result.report_info.stock_count}")
        print(f"   文件大小: {result.report_info.file_size} 字节")
        print(f"   历史数据行数: {len(df_history)}")
        
        # 清理测试文件
        os.remove(result.file_path)
    else:
        print("⚠️ 未生成报告文件（可能数据库中没有数据）")


def test_generate_report_with_stock_filter(report_service, test_user, test_watchlist):
    """测试生成指定股票的报告"""
    result = report_service.generate_user_report(
        test_user.id, 
        'summary', 
        stock_codes=["000001"]
    )
    
    assert result.success is True
    assert result.report_info.stock_count == 1
    
    if result.file_path:
        df = pd.read_csv(result.file_path, encoding='utf-8-sig')
        assert len(df) == 1
        # 股票代码可能是字符串或整数，统一转换为字符串比较
        assert str(df.iloc[0]['股票代码']) == "000001" or df.iloc[0]['股票代码'] == 1
        
        print(f"✅ 生成指定股票报告成功")
        
        # 清理测试文件
        os.remove(result.file_path)


def test_get_report_info(report_service, test_user, test_watchlist):
    """测试获取报告信息"""
    # 先生成一个报告
    result = report_service.generate_user_report(test_user.id, 'summary')
    
    if result.file_path:
        # 获取报告信息
        report_info = report_service.get_report_info(result.file_path)
        
        assert report_info is not None
        assert report_info.stock_count == 3
        assert report_info.report_type == 'summary'
        assert report_info.file_size > 0
        
        print(f"✅ 获取报告信息成功")
        print(f"   股票数量: {report_info.stock_count}")
        print(f"   报告类型: {report_info.report_type}")
        print(f"   文件大小: {report_info.file_size} 字节")
        
        # 清理测试文件
        os.remove(result.file_path)


def test_get_report_info_nonexistent_file(report_service):
    """测试获取不存在文件的报告信息"""
    report_info = report_service.get_report_info("nonexistent_file.csv")
    assert report_info is None


def test_invalid_report_type(report_service, test_user, test_watchlist):
    """测试无效的报告类型"""
    result = report_service.generate_user_report(test_user.id, 'invalid_type')
    
    assert result.success is False
    assert result.file_path is None
    assert "不支持的报告类型" in result.error_message


def test_volume_aberration_report_no_data(report_service, test_user):
    """测试成交量异动榜报告：无数据时返回成功、has_data=False"""
    from unittest.mock import patch
    from backend_api.services import volume_aberration_service
    with patch.object(volume_aberration_service, 'get_volume_aberration_data') as m:
        m.return_value = ([], None)
        result = report_service.generate_user_report(test_user.id, 'volume_aberration')
    assert result.success is True
    assert result.report_info is not None
    assert result.report_info.report_type == 'volume_aberration'
    assert result.report_info.has_data is False
    assert result.report_info.stock_count == 0
    assert result.file_path is None


def test_volume_aberration_report_with_data(report_service, test_user):
    """测试成交量异动榜报告：有数据时生成 Excel、report_type 与 has_data 正确"""
    import re
    from openpyxl import load_workbook
    from unittest.mock import patch
    from backend_api.services import volume_aberration_service
    row = {
        "rank": 1, "code": "000001", "name": "平安银行", "date": "2026-01-01",
        "volume": 1e7, "amount": 1e9, "mavol5": 8e6, "mavol10": 9e6, "mavol20": 9e6,
        "ratio_5": 1.25, "ratio_20": 3.0, "change_percent": 2.5, "close": 12.0, "turnover_rate": 5.0,
    }
    with patch.object(volume_aberration_service, 'get_volume_aberration_data', side_effect=[
        ([row], "2026-01-01"),  # cn
        ([], None),              # hk
    ]):
        result = report_service.generate_user_report(test_user.id, 'volume_aberration')
    assert result.success is True
    assert result.report_info.report_type == 'volume_aberration'
    assert result.report_info.has_data is True
    assert result.report_info.stock_count == 1
    assert result.file_path is not None
    assert result.file_path.endswith('.xlsx')
    assert re.search(
        rf"volume_aberration_{test_user.id}_20260101_\d{{6}}\.xlsx$",
        result.file_path.replace('\\', '/').split('/')[-1],
    ), f"unexpected filename: {result.file_path}"
    if os.path.exists(result.file_path):
        wb = load_workbook(result.file_path)
        ws = wb["A股放量榜"]
        # 校验成交额格式化：amount=1e9 -> 10.00亿
        header = [c.value for c in ws[1]]
        amount_col = header.index("成交额") + 1
        assert ws.cell(row=2, column=amount_col).value == "10.00亿"
        # 数据行从第2行开始；ratio_20 > 2.5 应触发行整行填充
        cell = ws.cell(row=2, column=1)  # A2
        assert cell.fill.patternType == "solid"
        # deep_fill = PatternFill(start_color="FFC00000", ...)
        assert cell.fill.start_color.rgb in ("FFC00000", "FF C00000".replace(" ", ""))
        os.remove(result.file_path)


if __name__ == "__main__":
    print("=" * 80)
    print("测试 ReportService - 报告生成服务")
    print("=" * 80)
    
    # 运行测试
    pytest.main([__file__, "-v", "-s"])
