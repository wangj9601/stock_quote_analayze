"""cn_listed_board_filter 单元测试。"""

import pytest

from backend_api.utils.cn_listed_board_filter import (
    classify_tvo_excel_board_segment,
    group_tvo_rows_by_excel_board,
    normalize_list_board_segment,
)
from backend_core.strategies.volume_shrink_breakout.data_loader import (
    code_matches_vsb_boards,
    normalize_vsb_board_keys,
)


def test_normalize_main_expands():
    assert normalize_list_board_segment("MAIN") == ["SH_MAIN", "SZ_MAIN"]
    assert normalize_list_board_segment("main") == ["SH_MAIN", "SZ_MAIN"]


def test_normalize_single_board():
    assert normalize_list_board_segment("CYB") == ["CYB"]
    assert normalize_list_board_segment("") == []
    assert normalize_list_board_segment(None) == []


def test_main_matches_prefixes():
    keys = normalize_vsb_board_keys(["SH_MAIN", "SZ_MAIN"])
    assert code_matches_vsb_boards("000981", keys)
    assert code_matches_vsb_boards("600000", keys)
    assert not code_matches_vsb_boards("300001", keys)
    assert not code_matches_vsb_boards("688001", keys)


def test_classify_tvo_excel_board_segment():
    assert classify_tvo_excel_board_segment("600000", "CN") == "MAIN"
    assert classify_tvo_excel_board_segment("000001", "CN") == "MAIN"
    assert classify_tvo_excel_board_segment("002001", "CN") == "SZ_SME"
    assert classify_tvo_excel_board_segment("300018", "CN") == "CYB"
    assert classify_tvo_excel_board_segment("688001", "CN") == "KCB"
    assert classify_tvo_excel_board_segment("00700", "HK") == "HK"


def test_group_tvo_rows_by_excel_board():
    rows = [
        {"市场": "CN", "代码": "600000", "名称": "A"},
        {"市场": "CN", "代码": "002001", "名称": "B"},
        {"市场": "CN", "代码": "300018", "名称": "C"},
        {"市场": "CN", "代码": "688001", "名称": "D"},
    ]
    buckets = group_tvo_rows_by_excel_board(rows)
    assert len(buckets["MAIN"]) == 1
    assert len(buckets["SZ_SME"]) == 1
    assert len(buckets["CYB"]) == 1
    assert len(buckets["KCB"]) == 1


def test_write_triple_volume_scan_push_excel_sheets(tmp_path):
    import pandas as pd
    from openpyxl import load_workbook

    from backend_api.services.report_service import ReportService

    path = tmp_path / "tvo.xlsx"
    rows = [
        {"市场": "CN", "代码": "600000", "名称": "浦发行", "观察日": "2024-01-02",
         "前日": "", "前日量": 1, "当日量": 2, "量比": 3, "状态": "待观察", "复核时间": "", "VSB摘要": ""},
        {"市场": "CN", "代码": "300018", "名称": "中元", "观察日": "2024-01-02",
         "前日": "", "前日量": 1, "当日量": 2, "量比": 3, "状态": "待观察", "复核时间": "", "VSB摘要": ""},
    ]
    ReportService._write_triple_volume_scan_push_excel(str(path), rows)
    wb = load_workbook(path)
    assert wb.sheetnames == ["沪深主板", "中小板", "创业板", "科创板"]
    assert wb["沪深主板"].max_row == 2
    assert wb["创业板"].max_row == 2
    assert wb["中小板"].max_row == 1
