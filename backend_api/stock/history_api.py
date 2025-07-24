import codecs
from fastapi import APIRouter, Query, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from backend_api.database import get_db
from typing import List, Optional
import io
import csv
from sqlalchemy import text
from datetime import datetime

# 新增依赖
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

router = APIRouter(prefix="/api/stock/history", tags=["StockHistory"])

def format_date_yyyymmdd(date_str: Optional[str]) -> Optional[str]:
    if not date_str:
        return None
    # 支持 "YYYY-MM-DD"、"YYYY/MM/DD"、"YYYY.MM.DD" 等
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    # 如果本身就是8位数字，尝试转为YYYY-MM-DD格式
    if len(date_str) == 8 and date_str.isdigit():
        try:
            return datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")
        except Exception:
            pass
    return date_str  # fallback

@router.get("")
def get_stock_history(
    code: str = Query(...),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    start_date_fmt = format_date_yyyymmdd(start_date)
    end_date_fmt = format_date_yyyymmdd(end_date)
    print(f"[get_stock_history] 输入参数: code={code}, start_date={start_date_fmt}, end_date={end_date_fmt}, page={page}, size={size}")
    query = "SELECT code,name,date, open, close, high, low, volume,amount,change_percent,change FROM historical_quotes WHERE code = :code"
    params = {"code": code}
    if start_date_fmt:
        query += " AND date >= :start_date"
        params["start_date"] = start_date_fmt
    if end_date_fmt:
        query += " AND date <= :end_date"
        params["end_date"] = end_date_fmt
    query += " ORDER BY date DESC"
    count_query = f"SELECT COUNT(*) FROM ({query})"
    total = db.execute(text(count_query), params).scalar()

    query += " LIMIT :limit OFFSET :offset"
    params["limit"] = size
    params["offset"] = (page - 1) * size
    result = db.execute(text(query), params)   
    items = [
        {
            "code": row[0],
            "name": row[1],
            "date": row[2],
            "open": row[3],
            "close": row[4],
            "high": row[5],
            "low": row[6],
            "volume": row[7],
            "amount": row[8],
            "change_percent": row[9],
            "change": row[10]
        }
        for row in result.fetchall()
    ]
    print(f"[get_stock_history] 输出: total={total}, items_count={len(items)}")
    return {"items": items, "total": total}

@router.get("/export")
def export_stock_history(
    code: str = Query(...),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    start_date_fmt = format_date_yyyymmdd(start_date)
    end_date_fmt = format_date_yyyymmdd(end_date)
    print(f"[export_stock_history] 输入参数: code={code}, start_date={start_date_fmt}, end_date={end_date_fmt}")
    # 字段顺序与表头一致
    query = "SELECT code, name, date, open, close, high, low, volume, amount, change_percent, change FROM historical_quotes WHERE code = :code"
    params = {"code": code}
    if start_date_fmt:
        query += " AND date >= :start_date"
        params["start_date"] = start_date_fmt
    if end_date_fmt:
        query += " AND date <= :end_date"
        params["end_date"] = end_date_fmt
    query += " ORDER BY date DESC"
    result = db.execute(text(query), params)
    rows = result.fetchall()
    columns = ["股票代码", "股票名称", "日期", "开盘价", "收盘价", "最高价", "最低价", "成交量", "成交额", "涨跌幅", "涨跌额"]
    df = pd.DataFrame(rows, columns=columns)

    df["成交量"] = df["成交量"].apply(lambda x: f"{round(float(x)/1e4, 2)}万" if x is not None and x != '' else '')
    df["成交额"] = df["成交额"].apply(lambda x: f"{round(float(x)/1e8, 2)}亿" if x is not None and x != '' else '')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="历史行情")
        worksheet = writer.sheets["历史行情"]
        # 标题加粗
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        # 列宽自适应
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = max_length + 2
        # 设置涨跌幅、涨跌额颜色
        from openpyxl.styles import Font as XLFont
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # "涨跌幅"在第10列（J），"涨跌额"在第11列（K），列索引从0开始
            change_percent_cell = row[9]
            change_cell = row[10]
            # 涨跌幅
            try:
                val = change_percent_cell.value
                if val is not None and val != '':
                    num = float(str(val).replace('%',''))
                    if num > 0:
                        change_percent_cell.font = XLFont(color='FF0000', bold=True)
                    elif num < 0:
                        change_percent_cell.font = XLFont(color='009933', bold=True)
            except:
                pass
            # 涨跌额
            try:
                val = change_cell.value
                if val is not None and val != '':
                    num = float(val)
                    if num > 0:
                        change_cell.font = XLFont(color='FF0000', bold=True)
                    elif num < 0:
                        change_cell.font = XLFont(color='009933', bold=True)
            except:
                pass
    output.seek(0)
    filename = f"{code}_history.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )