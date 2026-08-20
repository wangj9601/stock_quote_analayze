"""
报告服务 - 封装CSV报告生成器
"""

import os
import logging
from datetime import datetime, date
from typing import List, Optional, Dict, Any
from dataclasses import dataclass
import pandas as pd
from sqlalchemy.orm import Session
from sqlalchemy import text, func

from backend_api.models import (
    Watchlist,
    HistoricalQuotes,
    HistoricalQuotesHK,
    StockBasicInfo,
    StockBasicInfoHK,
    TripleVolumeObserveStock,
)

logger = logging.getLogger(__name__)


@dataclass
class ReportInfo:
    """报告信息"""
    stock_count: int  # 股票数量
    report_date: str  # 报告日期
    report_type: str  # 报告类型
    file_size: int  # 文件大小（字节）
    has_data: bool  # 是否有数据
    missing_data_stocks: List[str]  # 数据缺失的股票代码列表


@dataclass
class ReportResult:
    """报告生成结果"""
    success: bool  # 是否成功
    file_path: Optional[str]  # 报告文件路径
    report_info: Optional[ReportInfo]  # 报告信息
    error_message: Optional[str]  # 错误信息


class ReportService:
    """报告生成服务 - 封装CSV报告生成器"""
    
    def __init__(self, db: Session, report_dir: str = "reports/csv"):
        """
        初始化报告服务
        
        Args:
            db: 数据库会话
            report_dir: 报告文件保存目录
        """
        self.db = db
        self.report_dir = report_dir
        
        # 确保报告目录存在
        os.makedirs(self.report_dir, exist_ok=True)
        
        logger.info(f"ReportService 初始化完成，报告目录: {self.report_dir}")
    
    def get_user_watchlist(self, user_id: int, stock_codes: Optional[List[str]] = None) -> List[Dict[str, Any]]:
        """
        获取用户自选股列表
        
        Args:
            user_id: 用户ID
            stock_codes: 指定股票代码列表，None表示获取全部自选股
            
        Returns:
            自选股列表，每项包含 stock_code, stock_name, market
        """
        try:
            query = self.db.query(Watchlist).filter(Watchlist.user_id == user_id)
            
            # 如果指定了股票代码，则过滤
            if stock_codes:
                query = query.filter(Watchlist.stock_code.in_(stock_codes))
            
            watchlist_items = query.all()
            
            if not watchlist_items:
                logger.warning(f"用户 {user_id} 没有自选股数据")
                return []
            
            # 构建返回结果，需要判断市场类型
            result = []
            for item in watchlist_items:
                # 判断是A股还是港股（港股代码通常以数字开头且长度为5位）
                market = self._determine_market(item.stock_code)
                result.append({
                    'stock_code': item.stock_code,
                    'stock_name': item.stock_name,
                    'market': market
                })
            
            logger.info(f"获取用户 {user_id} 自选股 {len(result)} 只")
            return result
            
        except Exception as e:
            logger.error(f"获取用户自选股失败: {str(e)}")
            return []
    
    def _determine_market(self, stock_code: str) -> str:
        """
        判断股票市场类型
        
        Args:
            stock_code: 股票代码
            
        Returns:
            'CN' 表示A股，'HK' 表示港股
        """
        # 港股代码通常是5位数字，如 "00700"
        # A股代码通常是6位数字，如 "000001", "600000"
        if len(stock_code) == 5 and stock_code.isdigit():
            return 'HK'
        return 'CN'
    
    def get_stock_history_data(
        self, 
        stock_code: str, 
        market: str, 
        days: int = 30
    ) -> List[Dict[str, Any]]:
        """
        获取股票历史数据
        
        Args:
            stock_code: 股票代码
            market: 市场类型 ('CN' 或 'HK')
            days: 获取最近多少天的数据
            
        Returns:
            历史数据列表
        """
        try:
            if market == 'HK':
                # 查询港股历史数据
                query = text("""
                    SELECT 
                        date as trade_date,
                        open as open_price,
                        high as high_price,
                        low as low_price,
                        close as close_price,
                        volume,
                        amount,
                        change_amount,
                        change_percent
                    FROM historical_quotes_hk
                    WHERE code = :stock_code
                    ORDER BY date DESC
                    LIMIT :days
                """)
            else:
                # 查询A股历史数据
                query = text("""
                    SELECT 
                        date as trade_date,
                        open as open_price,
                        high as high_price,
                        low as low_price,
                        close as close_price,
                        volume,
                        amount,
                        change as change_amount,
                        change_percent
                    FROM historical_quotes
                    WHERE code = :stock_code
                    ORDER BY date DESC
                    LIMIT :days
                """)
            
            result = self.db.execute(query, {'stock_code': stock_code, 'days': days})
            rows = result.fetchall()
            
            # 转换为字典列表
            data = []
            for row in rows:
                data.append({
                    'trade_date': str(row.trade_date),
                    'open_price': float(row.open_price) if row.open_price else 0.0,
                    'high_price': float(row.high_price) if row.high_price else 0.0,
                    'low_price': float(row.low_price) if row.low_price else 0.0,
                    'close_price': float(row.close_price) if row.close_price else 0.0,
                    'volume': float(row.volume) if row.volume else 0.0,
                    'amount': float(row.amount) if row.amount else 0.0,
                    'change_amount': float(row.change_amount) if row.change_amount else 0.0,
                    'change_percent': float(row.change_percent) if row.change_percent else 0.0
                })
            
            if not data:
                logger.warning(f"股票 {stock_code} ({market}) 没有历史数据")
            
            return data
            
        except Exception as e:
            logger.error(f"获取股票 {stock_code} ({market}) 历史数据失败: {str(e)}")
            return []
    
    def get_stock_summary_data(self, stock_code: str, market: str) -> Dict[str, Any]:
        """
        获取股票汇总数据（最新一条数据）
        
        Args:
            stock_code: 股票代码
            market: 市场类型 ('CN' 或 'HK')
            
        Returns:
            汇总数据字典
        """
        try:
            if market == 'HK':
                # 查询港股最新数据
                query = text("""
                    SELECT 
                        h.name as stock_name,
                        h.close as current_price,
                        h.change_amount,
                        h.change_percent,
                        h.volume,
                        h.amount,
                        h.date as trade_date
                    FROM historical_quotes_hk h
                    WHERE h.code = :stock_code
                    ORDER BY h.date DESC
                    LIMIT 1
                """)
            else:
                # 查询A股最新数据
                query = text("""
                    SELECT 
                        h.name as stock_name,
                        h.close as current_price,
                        h.change as change_amount,
                        h.change_percent,
                        h.volume,
                        h.amount,
                        h.date as trade_date
                    FROM historical_quotes h
                    WHERE h.code = :stock_code
                    ORDER BY h.date DESC
                    LIMIT 1
                """)
            
            result = self.db.execute(query, {'stock_code': stock_code})
            row = result.fetchone()
            
            if not row:
                logger.warning(f"股票 {stock_code} ({market}) 没有汇总数据")
                return {}
            
            return {
                'stock_name': row.stock_name,
                'market': market,
                'current_price': float(row.current_price) if row.current_price else 0.0,
                'change_amount': float(row.change_amount) if row.change_amount else 0.0,
                'change_percent': float(row.change_percent) if row.change_percent else 0.0,
                'volume': float(row.volume) if row.volume else 0.0,
                'amount': float(row.amount) if row.amount else 0.0,
                'trade_date': str(row.trade_date)
            }
            
        except Exception as e:
            logger.error(f"获取股票 {stock_code} ({market}) 汇总数据失败: {str(e)}")
            return {}
    
    def generate_user_report(
        self, 
        user_id: int, 
        report_type: str,
        stock_codes: Optional[List[str]] = None
    ) -> ReportResult:
        """
        生成用户报告
        
        Args:
            user_id: 用户ID
            report_type: 报告类型 ('summary' 或 'detailed')
            stock_codes: 指定股票代码列表，None表示全部自选股
            
        Returns:
            ReportResult: 报告生成结果
        """
        try:
            # 成交量异动榜不依赖自选股，直接生成
            if report_type == 'volume_aberration':
                return self._generate_volume_aberration_report(user_id)

            if report_type in ('triple_volume_observe_scan', 'triple_volume_observe_eval'):
                return self._generate_triple_volume_observe_report(user_id, report_type)

            # 获取用户自选股列表
            watchlist = self.get_user_watchlist(user_id, stock_codes)
            
            if not watchlist:
                logger.warning(f"用户 {user_id} 没有自选股，生成空报告")
                return ReportResult(
                    success=True,
                    file_path=None,
                    report_info=ReportInfo(
                        stock_count=0,
                        report_date=datetime.now().strftime("%Y-%m-%d"),
                        report_type=report_type,
                        file_size=0,
                        has_data=False,
                        missing_data_stocks=[]
                    ),
                    error_message="用户没有自选股"
                )
            
            # 根据报告类型生成报告
            if report_type == 'summary':
                return self._generate_summary_report(user_id, watchlist)
            elif report_type == 'detailed':
                return self._generate_detailed_report(user_id, watchlist)
            elif report_type == 'gms_daily':
                return self._generate_gms_report_for_user(user_id)
            elif report_type == 'urt_daily':
                return self._generate_urt_report_for_user(user_id, stock_codes=stock_codes)
            else:
                return ReportResult(
                    success=False,
                    file_path=None,
                    report_info=None,
                    error_message=f"不支持的报告类型: {report_type}"
                )
                
        except Exception as e:
            logger.error(f"生成用户 {user_id} 报告失败: {str(e)}")
            return ReportResult(
                success=False,
                file_path=None,
                report_info=None,
                error_message=str(e)
            )
    
    def _generate_summary_report(
        self, 
        user_id: int, 
        watchlist: List[Dict[str, Any]]
    ) -> ReportResult:
        """
        生成汇总报告（仅包含最新数据）
        
        Args:
            user_id: 用户ID
            watchlist: 自选股列表
            
        Returns:
            ReportResult: 报告生成结果
        """
        summary_data = []
        missing_data_stocks = []
        
        for stock in watchlist:
            stock_code = stock['stock_code']
            stock_name = stock['stock_name']
            market = stock['market']
            
            # 获取汇总数据
            summary = self.get_stock_summary_data(stock_code, market)
            
            if summary:
                summary_data.append({
                    '股票代码': stock_code,
                    '股票名称': stock_name,
                    '市场': market,
                    '当前价格': summary.get('current_price', 0),
                    '涨跌额': summary.get('change_amount', 0),
                    '涨跌幅(%)': summary.get('change_percent', 0),
                    '成交量': summary.get('volume', 0),
                    '成交额': summary.get('amount', 0),
                    '最新交易日': summary.get('trade_date', '')
                })
            else:
                # 数据缺失，添加占位行
                missing_data_stocks.append(stock_code)
                summary_data.append({
                    '股票代码': stock_code,
                    '股票名称': stock_name,
                    '市场': market,
                    '当前价格': '数据缺失',
                    '涨跌额': '数据缺失',
                    '涨跌幅(%)': '数据缺失',
                    '成交量': '数据缺失',
                    '成交额': '数据缺失',
                    '最新交易日': '数据缺失'
                })
        
        # 生成CSV文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"stock_summary_{user_id}_{timestamp}.csv"
        filepath = os.path.join(self.report_dir, filename)
        
        df = pd.DataFrame(summary_data)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        # 获取文件大小
        file_size = os.path.getsize(filepath)
        
        logger.info(f"生成汇总报告成功: {filepath}, 股票数: {len(watchlist)}, 数据缺失: {len(missing_data_stocks)}")
        
        return ReportResult(
            success=True,
            file_path=filepath,
            report_info=ReportInfo(
                stock_count=len(watchlist),
                report_date=datetime.now().strftime("%Y-%m-%d"),
                report_type='summary',
                file_size=file_size,
                has_data=len(summary_data) > 0,
                missing_data_stocks=missing_data_stocks
            ),
            error_message=None
        )
    
    def _generate_detailed_report(
        self, 
        user_id: int, 
        watchlist: List[Dict[str, Any]],
        days: int = 30
    ) -> ReportResult:
        """
        生成详细报告（包含历史数据）
        
        Args:
            user_id: 用户ID
            watchlist: 自选股列表
            days: 获取最近多少天的数据
            
        Returns:
            ReportResult: 报告生成结果
        """
        all_data = []
        summary_data = []
        missing_data_stocks = []
        
        for stock in watchlist:
            stock_code = stock['stock_code']
            stock_name = stock['stock_name']
            market = stock['market']
            
            # 获取历史数据
            history_data = self.get_stock_history_data(stock_code, market, days)
            
            # 获取汇总数据
            summary = self.get_stock_summary_data(stock_code, market)
            
            # 处理历史数据
            if history_data:
                for data in history_data:
                    all_data.append({
                        '股票代码': stock_code,
                        '股票名称': stock_name,
                        '市场': market,
                        '交易日期': data['trade_date'],
                        '开盘价': data['open_price'],
                        '最高价': data['high_price'],
                        '最低价': data['low_price'],
                        '收盘价': data['close_price'],
                        '成交量': data['volume'],
                        '成交额': data['amount'],
                        '涨跌额': data['change_amount'],
                        '涨跌幅(%)': data['change_percent']
                    })
            else:
                # 历史数据缺失
                missing_data_stocks.append(stock_code)
                all_data.append({
                    '股票代码': stock_code,
                    '股票名称': stock_name,
                    '市场': market,
                    '交易日期': '数据缺失',
                    '开盘价': '数据缺失',
                    '最高价': '数据缺失',
                    '最低价': '数据缺失',
                    '收盘价': '数据缺失',
                    '成交量': '数据缺失',
                    '成交额': '数据缺失',
                    '涨跌额': '数据缺失',
                    '涨跌幅(%)': '数据缺失'
                })
            
            # 处理汇总数据
            if summary:
                summary_data.append({
                    '股票代码': stock_code,
                    '股票名称': stock_name,
                    '市场': market,
                    '当前价格': summary.get('current_price', 0),
                    '涨跌额': summary.get('change_amount', 0),
                    '涨跌幅(%)': summary.get('change_percent', 0),
                    '成交量': summary.get('volume', 0),
                    '成交额': summary.get('amount', 0),
                    '最新交易日': summary.get('trade_date', '')
                })
            else:
                summary_data.append({
                    '股票代码': stock_code,
                    '股票名称': stock_name,
                    '市场': market,
                    '当前价格': '数据缺失',
                    '涨跌额': '数据缺失',
                    '涨跌幅(%)': '数据缺失',
                    '成交量': '数据缺失',
                    '成交额': '数据缺失',
                    '最新交易日': '数据缺失'
                })
        
        # 生成Excel文件，包含多个工作表
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"stock_report_{user_id}_{timestamp}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        
        # 创建Excel文件，包含多个工作表
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 历史数据表
            if all_data:
                df_history = pd.DataFrame(all_data)
                df_history.to_excel(writer, sheet_name='历史数据', index=False)
            
            # 汇总数据表
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='股票汇总', index=False)
        
        # 获取文件大小
        file_size = os.path.getsize(filepath)
        
        logger.info(f"生成详细报告成功: {filepath}, 股票数: {len(watchlist)}, 数据缺失: {len(missing_data_stocks)}")
        
        return ReportResult(
            success=True,
            file_path=filepath,
            report_info=ReportInfo(
                stock_count=len(watchlist),
                report_date=datetime.now().strftime("%Y-%m-%d"),
                report_type='detailed',
                file_size=file_size,
                has_data=len(all_data) > 0,
                missing_data_stocks=missing_data_stocks
            ),
            error_message=None
        )

    def _generate_volume_aberration_report(self, user_id: int) -> ReportResult:
        """
        生成成交量异动榜报告（A股+港股放量榜全量），不依赖自选股。
        输出 Excel 两 sheet：A股放量榜、港股放量榜。
        """
        from backend_api.services.volume_aberration_service import get_volume_aberration_data

        result_cn, date_cn = get_volume_aberration_data(self.db, market="cn", date=None, order="desc")
        result_hk, date_hk = get_volume_aberration_data(self.db, market="hk", date=None, order="desc")

        report_date = date_cn or date_hk or datetime.now().strftime("%Y-%m-%d")
        total_count = len(result_cn) + len(result_hk)

        if total_count == 0:
            return ReportResult(
                success=True,
                file_path=None,
                report_info=ReportInfo(
                    stock_count=0,
                    report_date=report_date,
                    report_type="volume_aberration",
                    file_size=0,
                    has_data=False,
                    missing_data_stocks=[],
                ),
                error_message=None,
            )

        def to_rows(data: List[Dict[str, Any]], default_trade_date: Optional[str]) -> List[Dict[str, Any]]:
            def format_volume_hand(v: Any) -> str:
                """成交量/均量展示为手口径数值（不显示“手”字）。"""
                if v is None or v == "":
                    return ""
                try:
                    num = float(v)
                except (ValueError, TypeError):
                    return str(v)

                if num >= 100000000:
                    return f"{(num / 100000000):.2f}亿"
                if num >= 10000:
                    return f"{(num / 10000):.2f}万"
                return f"{num:.0f}"

            def format_amount(v: Any) -> str:
                """成交额按“万/亿”展示（最小单位为万）。"""
                if v is None or v == "":
                    return ""
                try:
                    num = float(v)
                except (ValueError, TypeError):
                    return str(v)

                if num >= 100000000:
                    return f"{(num / 100000000):.2f}亿"
                # 最小单位为“万”，即便小于 1 万也显示为 0.xx 万
                return f"{(num / 10000):.2f}万"

            def format_two_decimals(v: Any) -> Any:
                """涨跌幅、量比等：保留两位小数；空值保持空。"""
                if v is None or v == "":
                    return ""
                try:
                    return f"{float(v):.2f}"
                except (ValueError, TypeError):
                    return str(v)

            rows = []
            for item in data:
                # 日期列以历史行情交易日为准；若行内缺失，回退到该市场实际交易日（date_cn/date_hk），不使用运行当天
                trade_date = (str(item.get("date", "")).strip()[:10]) or (default_trade_date or "")
                rows.append({
                    "排名": item.get("rank"),
                    "股票代码": "\u2060" + str(item.get("code", "")),
                    "股票名称": (item.get("name") or "").strip(),
                    "日期": trade_date,
                    "当日成交量(手)": format_volume_hand(item.get("volume")),
                    "成交额(元)": format_amount(item.get("amount")),
                    "MAVOL5(手)": format_volume_hand(item.get("mavol5")),
                    "MAVOL10(手)": format_volume_hand(item.get("mavol10")),
                    "MAVOL20(手)": format_volume_hand(item.get("mavol20")),
                    "量比(5)": format_two_decimals(item.get("ratio_5")),
                    "量比(20)": format_two_decimals(item.get("ratio_20")),
                    "涨跌幅(%)": format_two_decimals(item.get("change_percent")),
                    "收盘价": item.get("close"),
                    "换手率(%)": item.get("turnover_rate"),
                })
            return rows

        rows_cn = to_rows(result_cn, date_cn)
        rows_hk = to_rows(result_hk, date_hk)

        # 文件名里 report_date 已包含日期信息，避免 timestamp 再附带日期导致重复
        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"volume_aberration_{user_id}_{report_date.replace('-', '')}_{timestamp}.xlsx"
        filepath = os.path.join(self.report_dir, filename)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            if rows_cn:
                pd.DataFrame(rows_cn).to_excel(writer, sheet_name="A股放量榜", index=False)
            if rows_hk:
                pd.DataFrame(rows_hk).to_excel(writer, sheet_name="港股放量榜", index=False)

        # 量比(20) > 2.5 时，仅「量比(20)」列单元格使用红色字体（不整行着色、不设底色）
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            wb = load_workbook(filepath)

            def apply_volume_aberration_column_widths(ws):
                """按表头名加宽常用列，便于阅读（股票名称、日期、成交量、成交额、MAVOL 等）。"""
                if ws.max_row < 1:
                    return
                title_to_col = {}
                for cell in ws[1]:
                    if cell.value is not None:
                        title_to_col[str(cell.value).strip()] = cell.column
                # 单位：Excel 列宽约等于字符数；中文列略加宽
                desired = {
                    "股票名称": 20,
                    "日期": 13,
                    "当日成交量(手)": 22,
                    "成交额(元)": 18,
                    "MAVOL5(手)": 16,
                    "MAVOL10(手)": 16,
                    "MAVOL20(手)": 16,
                    "涨跌幅(%)": 14,
                    "量比(5)": 12,
                    "量比(20)": 12,
                }
                for title, width in desired.items():
                    col_idx = title_to_col.get(title)
                    if col_idx:
                        ws.column_dimensions[get_column_letter(col_idx)].width = width

            for _sn in ("A股放量榜", "港股放量榜"):
                if _sn in wb.sheetnames:
                    apply_volume_aberration_column_widths(wb[_sn])

            ratio_highlight_font = Font(color="FFFF0000")  # 标准红字

            def apply_sheet_style(sheet_name: str, rows_data: List[Dict[str, Any]]):
                if sheet_name not in wb.sheetnames or not rows_data:
                    return

                ws = wb[sheet_name]
                # 表头行=1，找到“量比(20)”列
                ratio_col = None
                for cell in ws[1]:
                    if cell.value == "量比(20)":
                        ratio_col = cell.column
                        break
                if not ratio_col:
                    return

                for idx, item in enumerate(rows_data):
                    ratio_val = item.get("量比(20)")
                    if ratio_val is None:
                        continue
                    try:
                        ratio_f = float(ratio_val)
                    except (ValueError, TypeError):
                        continue
                    if ratio_f > 2.5:
                        row_idx = 2 + idx  # 数据从第2行开始
                        c = ws.cell(row=row_idx, column=ratio_col)
                        c.font = ratio_highlight_font

            apply_sheet_style("A股放量榜", rows_cn)
            apply_sheet_style("港股放量榜", rows_hk)

            wb.save(filepath)
        except Exception as e:
            # 样式失败不应影响报告生成成功
            logger.warning("成交量异动榜 Excel 样式应用失败: %s", e)

        file_size = os.path.getsize(filepath)
        logger.info("生成成交量异动榜报告成功: %s, A股=%s, 港股=%s", filepath, len(rows_cn), len(rows_hk))

        return ReportResult(
            success=True,
            file_path=filepath,
            report_info=ReportInfo(
                stock_count=total_count,
                report_date=report_date,
                report_type="volume_aberration",
                file_size=file_size,
                has_data=True,
                missing_data_stocks=[],
            ),
            error_message=None,
        )

    @staticmethod
    def _triple_volume_observe_row_dict(r: TripleVolumeObserveStock) -> Dict[str, Any]:
        vd = r.vsb_detail_json if isinstance(r.vsb_detail_json, dict) else {}
        return {
            "市场": r.market,
            "代码": r.code,
            "名称": r.name or "",
            "观察日": r.observe_trade_date.strftime("%Y-%m-%d")
            if hasattr(r.observe_trade_date, "strftime")
            else str(r.observe_trade_date)[:10],
            "前日": r.prev_trade_date.strftime("%Y-%m-%d")
            if r.prev_trade_date and hasattr(r.prev_trade_date, "strftime")
            else ("" if not r.prev_trade_date else str(r.prev_trade_date)[:10]),
            "前日量": r.prev_volume,
            "当日量": r.curr_volume,
            "量比": round(r.volume_ratio_actual, 4) if r.volume_ratio_actual is not None else "",
            "状态": r.status,
            "复核时间": r.vsb_evaluated_at.strftime("%Y-%m-%d %H:%M:%S") if r.vsb_evaluated_at else "",
            "VSB摘要": str(vd) if vd else "",
        }

    @staticmethod
    def _write_triple_volume_scan_push_excel(filepath: str, rows: List[Dict[str, Any]]) -> None:
        """每日爆量推送 Excel：沪深主板 / 中小板 / 创业板 / 科创板 各一 sheet；港股单独 sheet。"""
        from backend_api.utils.cn_listed_board_filter import (
            TVO_PUSH_EXCEL_BOARD_SHEETS,
            TVO_PUSH_EXCEL_HK_SHEET,
            group_tvo_rows_by_excel_board,
        )

        column_order = [
            "市场",
            "代码",
            "名称",
            "观察日",
            "前日",
            "前日量",
            "当日量",
            "量比",
            "状态",
            "复核时间",
            "VSB摘要",
        ]
        buckets = group_tvo_rows_by_excel_board(rows)
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            for seg, sheet_name in TVO_PUSH_EXCEL_BOARD_SHEETS:
                seg_rows = buckets.get(seg, [])
                df = (
                    pd.DataFrame(seg_rows, columns=column_order)
                    if seg_rows
                    else pd.DataFrame(columns=column_order)
                )
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            hk_rows = buckets.get("HK", [])
            if hk_rows:
                pd.DataFrame(hk_rows, columns=column_order).to_excel(
                    writer, sheet_name=TVO_PUSH_EXCEL_HK_SHEET, index=False
                )
            other_rows = buckets.get("OTHER", [])
            if other_rows:
                pd.DataFrame(other_rows, columns=column_order).to_excel(
                    writer, sheet_name="其他", index=False
                )

    def _generate_triple_volume_observe_report(self, user_id: int, report_type: str) -> ReportResult:
        """3倍量观察股：scan=仅最新观察日；eval=全表快照。"""
        q = self.db.query(TripleVolumeObserveStock)
        if report_type == "triple_volume_observe_scan":
            md = self.db.query(func.max(TripleVolumeObserveStock.observe_trade_date)).scalar()
            if md is not None:
                q = q.filter(TripleVolumeObserveStock.observe_trade_date == md)
        rows = q.order_by(
            TripleVolumeObserveStock.observe_trade_date.desc(),
            TripleVolumeObserveStock.market,
            TripleVolumeObserveStock.code,
        ).all()

        report_date = datetime.now().strftime("%Y-%m-%d")
        if rows:
            d0 = rows[0].observe_trade_date
            report_date = d0.strftime("%Y-%m-%d") if hasattr(d0, "strftime") else str(d0)[:10]

        data = [self._triple_volume_observe_row_dict(r) for r in rows]

        timestamp = datetime.now().strftime("%H%M%S")
        short = "tvo_scan" if report_type == "triple_volume_observe_scan" else "tvo_eval"
        filename = f"{short}_{user_id}_{report_date.replace('-', '')}_{timestamp}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        if report_type == "triple_volume_observe_scan":
            self._write_triple_volume_scan_push_excel(filepath, data)
        else:
            df = pd.DataFrame(data)
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="状态复核", index=False)

        file_size = os.path.getsize(filepath)
        return ReportResult(
            success=True,
            file_path=filepath,
            report_info=ReportInfo(
                stock_count=len(rows),
                report_date=report_date,
                report_type=report_type,
                file_size=file_size,
                has_data=True,
                missing_data_stocks=[],
            ),
            error_message=None,
        )

    def _generate_gms_report_for_user(self, user_id: int) -> ReportResult:
        """
        生成该用户自选股范围内的 GMS 均值引力策略选股报告。
        """
        watchlist = self.get_user_watchlist(user_id)
        if not watchlist:
            return ReportResult(
                success=True,
                file_path=None,
                report_info=ReportInfo(
                    stock_count=0,
                    report_date=datetime.now().strftime("%Y-%m-%d"),
                    report_type="gms_daily",
                    file_size=0,
                    has_data=False,
                    missing_data_stocks=[],
                ),
                error_message="用户没有自选股",
            )
        # 自选股代码规范化：A 股 6 位，港股 5 位，与 GMS 指标表一致
        def _norm(s: str, market: str) -> str:
            s = str(s).strip()
            if not s:
                return s
            if market == "HK":
                return s.zfill(5) if s.isdigit() else s
            return s.zfill(6) if s.isdigit() else s

        stock_pool = []
        code_to_name = {}
        for s in watchlist:
            code = _norm(s["stock_code"], s["market"])
            if code:
                stock_pool.append(code)
                code_to_name[code] = s.get("stock_name") or ""

        if not stock_pool:
            return ReportResult(
                success=True,
                file_path=None,
                report_info=ReportInfo(
                    stock_count=0,
                    report_date=datetime.now().strftime("%Y-%m-%d"),
                    report_type="gms_daily",
                    file_size=0,
                    has_data=False,
                    missing_data_stocks=[],
                ),
                error_message="自选股代码无效",
            )

        try:
            from backend_core.strategies.gms.frontend_interface import GMSFrontendInterface

            gms = GMSFrontendInterface(self.db)
            results = gms.get_selection_results(date=None, stock_pool=stock_pool, market="all")
        except Exception as e:
            logger.exception("GMS 选股失败: %s", e)
            return ReportResult(
                success=False,
                file_path=None,
                report_info=None,
                error_message=str(e),
            )

        if not results:
            report_date = datetime.now().strftime("%Y-%m-%d")
            return ReportResult(
                success=True,
                file_path=None,
                report_info=ReportInfo(
                    stock_count=0,
                    report_date=report_date,
                    report_type="gms_daily",
                    file_size=0,
                    has_data=False,
                    missing_data_stocks=[],
                ),
                error_message="当日自选股范围内无 GMS 选股结果",
            )

        # 报告日期：优先使用 GMS 结果中的 date 字段（YYYY-MM-DD），否则为今天
        report_date = str(results[0].get("date", "")[:10]) if results else datetime.now().strftime("%Y-%m-%d")

        codes = [
            str(r.get("code") or r.get("symbol") or "").strip()
            for r in results
        ]
        cn_codes = [c for c in codes if c and len(c) >= 6 and c.isdigit() and c[0] in "6039"]
        hk_codes = [c for c in codes if c and c not in cn_codes]

        # 与前端选股一致：从历史行情表取最近交易日收盘价、涨跌幅，用于「当前价格」「当前涨跌幅」
        hist_quotes_a: Dict[str, Any] = {}
        hist_quotes_hk: Dict[str, Any] = {}
        if cn_codes:
            latest_date_a = self.db.query(func.max(HistoricalQuotes.date)).scalar()
            if latest_date_a:
                quotes_a = self.db.query(HistoricalQuotes).filter(
                    HistoricalQuotes.code.in_(cn_codes),
                    HistoricalQuotes.date == latest_date_a,
                ).all()
                hist_quotes_a = {q.code: q for q in quotes_a}
        if hk_codes:
            latest_date_hk = self.db.query(func.max(HistoricalQuotesHK.date)).scalar()
            if latest_date_hk:
                latest_hk_str = str(latest_date_hk).strip()[:10]
                quotes_hk = self.db.query(HistoricalQuotesHK).filter(
                    HistoricalQuotesHK.code.in_(hk_codes),
                    HistoricalQuotesHK.date == latest_hk_str,
                ).all()
                hist_quotes_hk = {q.code: q for q in quotes_hk}

        # 补全股票名称：先用自选股 code_to_name，缺失时查 StockBasicInfo / StockBasicInfoHK（与前端 API 一致）
        for code in codes:
            if code_to_name.get(code):
                continue
            if code in hk_codes:
                info = self.db.query(StockBasicInfoHK).filter(StockBasicInfoHK.code == code).first()
                if info and info.name:
                    code_to_name[code] = info.name
            else:
                info = self.db.query(StockBasicInfo).filter(StockBasicInfo.code == code).first()
                if not info and code.startswith("SZ"):
                    info = self.db.query(StockBasicInfo).filter(StockBasicInfo.code == code[2:]).first()
                if not info and code.startswith("SH"):
                    info = self.db.query(StockBasicInfo).filter(StockBasicInfo.code == code[2:]).first()
                if info and info.name:
                    code_to_name[code] = info.name
            if not code_to_name.get(code):
                code_to_name[code] = f"股票{code}"

        # CSV 表头与内容与前端「选股 - GMS均值引力动量」导出一致（见 frontend/js/screening.js strategy===gms）
        def _fmt_pct(val, decimals=2):
            if val is None:
                return ""
            try:
                return f"{float(val) * 100:.{decimals}f}%"
            except (TypeError, ValueError):
                return ""

        def _fmt_score_detail(sd):
            """与前端 screening.js GMS 导出格式一致：总分 蓄势(引力+平衡+量缩)等级 动量(推力+支撑+攻击)等级"""
            if not sd or not isinstance(sd, dict):
                return ""
            def _n(v):
                if v is None:
                    return "--"
                try:
                    return f"{float(v):.1f}"
                except (TypeError, ValueError):
                    return "--"
            acc = sd.get("score_accumulation")
            acc_part = f"蓄势{_n(acc)}(引力{_n(sd.get('score_acc_fz'))}+平衡{_n(sd.get('score_acc_balance'))}+量缩{_n(sd.get('score_acc_volume'))}){sd.get('accumulation_grade') or ''}" if acc is not None else "蓄势--"
            mom = sd.get("score_momentum")
            mom_part = f"动量{_n(mom)}(推力{_n(sd.get('score_mom_ratio_d1'))}+支撑{_n(sd.get('score_mom_deviation'))}+攻击{_n(sd.get('score_mom_volume'))}){sd.get('momentum_grade') or ''}" if mom is not None else "动量--"
            return f"总分{_n(sd.get('score_total'))} {acc_part} {mom_part}"

        rows = []
        for r in results:
            code = r.get("code") or r.get("symbol") or ""
            name = code_to_name.get(code, "")
            st = r.get("score_total")
            sig = (float(st) / 100.0) if st is not None and st > 0 else (r.get("signal_strength") or 0.0)
            buy_type = r.get("buy_type") or ""
            delta = r.get("delta")
            d = r.get("d")
            ratio_d20 = r.get("ratio_d20")
            ratio_d1 = r.get("ratio_d1")
            fz_ratio = r.get("fz_ratio")
            rising_days = r.get("rising_days")
            falling_days = r.get("falling_days")
            ratio_relative = (delta / d) if (delta is not None and d is not None and d != 0) else None

            quote = hist_quotes_a.get(code) or hist_quotes_hk.get(code)
            current_price = d or 0
            current_change_percent = None
            if quote and hasattr(quote, "close") and quote.close is not None:
                current_price = float(quote.close)
            if quote and hasattr(quote, "change_percent") and quote.change_percent is not None:
                current_change_percent = float(quote.change_percent)

            score_detail_str = _fmt_score_detail(r.get("score_detail"))

            rows.append({
                "股票代码": "\u2060" + str(code),  # 零宽字符前缀使 Excel 整列统一按文本显示，左对齐且保留前导零
                    "股票名称": name,
                "信号强度": f"{sig * 100:.1f}%",
                "买点类型": buy_type,
                "当前价格": f"{current_price:.2f}" if current_price is not None else "",
                "Δ (20日位移)": f"{delta:.2f}" if delta is not None else "",
                "F (下跌天)": falling_days if falling_days is not None else "",
                "Z (上涨天)": rising_days if rising_days is not None else "",
                "d (20日均价)": f"{d:.2f}" if d is not None else "",
                "Δ/d (位移/均价)": _fmt_pct(ratio_relative),
                "Δ/d₂₀": _fmt_pct(ratio_d20),
                "Δ/d₁": _fmt_pct(ratio_d1),
                "F/Z": f"{fz_ratio:.2f}" if fz_ratio is not None else "",
                "当前涨跌幅": f"{current_change_percent:.2f}%" if current_change_percent is not None else "0%",
                "得分明细": score_detail_str,
            })

        filename = f"gms_{user_id}_{report_date.replace('-', '')}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        df = pd.DataFrame(rows)
        if "股票代码" in df.columns:
            df["股票代码"] = df["股票代码"].astype(str)
        # GMS 推送报告改为 Excel 格式
        df.to_excel(filepath, index=False, sheet_name="GMS策略信号列表")
        # 设置 GMS 报告列宽，提升可读性
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(filepath)
            ws = wb["GMS策略信号列表"] if "GMS策略信号列表" in wb.sheetnames else wb[wb.sheetnames[0]]
            title_to_col = {}
            for cell in ws[1]:
                if cell.value is not None:
                    title_to_col[str(cell.value).strip()] = cell.column

            desired = {
                "股票代码": 12,
                "股票名称": 18,
                "信号强度": 12,
                "买点类型": 10,
                "当前价格": 12,
                "Δ (20日位移)": 14,
                "F (下跌天)": 10,
                "Z (上涨天)": 10,
                "d (20日均价)": 12,
                "Δ/d (位移/均价)": 14,
                "Δ/d₂₀": 12,
                "Δ/d₁": 12,
                "F/Z": 10,
                "当前涨跌幅": 12,
                "得分明细": 90,
            }
            for title, width in desired.items():
                col_idx = title_to_col.get(title)
                if col_idx:
                    ws.column_dimensions[get_column_letter(col_idx)].width = width

            wb.save(filepath)
        except Exception as e:
            logger.warning("GMS 报告 Excel 列宽样式应用失败: %s", e)
        file_size = os.path.getsize(filepath)
        logger.info("生成 GMS 自选股报告成功: %s, 选股数: %s", filepath, len(rows))
        return ReportResult(
            success=True,
            file_path=filepath,
            report_info=ReportInfo(
                stock_count=len(rows),
                report_date=report_date,
                report_type="gms_daily",
                file_size=file_size,
                has_data=True,
                missing_data_stocks=[],
            ),
            error_message=None,
        )

    @staticmethod
    def _urt_row_meets_report_filter(row: Dict[str, Any]) -> bool:
        """
        推送 Excel 收录规则（与 URT 连阳硬筛一致，不另行放宽）：
        - 正式买点（buy_signal=True）一律收录；
        - 非买点自选股：4日阳≥3 或 5日阳≥4（默认 yang_rule_a / yang_rule_b）。
        """
        if bool(row.get("buy_signal")):
            return True
        # 优先用策略已算好的规则结果
        if row.get("rule_a_ok") is True or row.get("rule_b_ok") is True:
            return True
        if row.get("rule_a_ok") is False and row.get("rule_b_ok") is False:
            return False
        try:
            y4 = int(row.get("yang_count_4") or 0)
        except (TypeError, ValueError):
            y4 = 0
        try:
            y5 = int(row.get("yang_count_5") or 0)
        except (TypeError, ValueError):
            y5 = 0
        # 与 indicators 默认一致：规则A window=4 min=3；规则B window=5 min=4
        return y4 >= 3 or y5 >= 4

    @staticmethod
    def _urt_pick_structure(row: Dict[str, Any]) -> Dict[str, Any]:
        """从选股结果行提取 structure 子对象。"""
        sd = row.get("score_detail") if isinstance(row.get("score_detail"), dict) else {}
        st = sd.get("structure") if isinstance(sd.get("structure"), dict) else {}
        return st if isinstance(st, dict) else {}

    @staticmethod
    def _urt_pick_support(row: Dict[str, Any]) -> Any:
        st = ReportService._urt_pick_structure(row)
        if row.get("nearest_support") is not None:
            return row.get("nearest_support")
        levels = row.get("support_levels")
        if isinstance(levels, list) and levels:
            return levels[0]
        if st.get("nearest_support") is not None:
            return st.get("nearest_support")
        st_levels = st.get("support_levels")
        if isinstance(st_levels, list) and st_levels:
            return st_levels[0]
        return None

    @staticmethod
    def _urt_pick_resistance(row: Dict[str, Any]) -> Any:
        st = ReportService._urt_pick_structure(row)
        if row.get("nearest_resistance") is not None:
            return row.get("nearest_resistance")
        levels = row.get("resistance_levels")
        if isinstance(levels, list) and levels:
            return levels[0]
        if st.get("nearest_resistance") is not None:
            return st.get("nearest_resistance")
        st_levels = st.get("resistance_levels")
        if isinstance(st_levels, list) and st_levels:
            return st_levels[0]
        return None

    @staticmethod
    def _urt_pick_structure_rr(row: Dict[str, Any]) -> Any:
        st = ReportService._urt_pick_structure(row)
        if row.get("structure_rr") is not None:
            return row.get("structure_rr")
        if st.get("rr") is not None:
            return st.get("rr")
        return None

    @staticmethod
    def _urt_format_risk_tags(row: Dict[str, Any]) -> str:
        sd = row.get("score_detail") if isinstance(row.get("score_detail"), dict) else {}
        tags = row.get("risk_tags") if isinstance(row.get("risk_tags"), list) else None
        if not tags:
            tags = sd.get("risk_tags") if isinstance(sd.get("risk_tags"), list) else []
        labels: List[str] = []
        for t in tags or []:
            if isinstance(t, dict):
                lab = str(t.get("label") or "").strip()
                if lab:
                    labels.append(lab)
            elif t:
                labels.append(str(t).strip())
        # 去重保序
        seen = set()
        out: List[str] = []
        for lab in labels:
            if lab in seen:
                continue
            seen.add(lab)
            out.append(lab)
        return "；".join(out)

    @staticmethod
    def _urt_format_trade_advice(row: Dict[str, Any]) -> str:
        sd = row.get("score_detail") if isinstance(row.get("score_detail"), dict) else {}
        advice = row.get("trade_advice") if isinstance(row.get("trade_advice"), dict) else None
        if not advice:
            advice = sd.get("trade_advice") if isinstance(sd.get("trade_advice"), dict) else {}
        if not isinstance(advice, dict) or not advice:
            return ""
        parts: List[str] = []
        action = str(advice.get("action") or "").strip()
        if action:
            parts.append(action)
        summary = str(advice.get("summary") or "").strip()
        if summary:
            parts.append(summary[:120])
        return "；".join(parts)

    @staticmethod
    def _urt_report_field_legend_rows() -> List[Dict[str, str]]:
        """
        URT 推送 Excel「字段说明」sheet 行数据。
        与 indicators / scoring / signal_detector / 选股页 / 本文件收录过滤逻辑一致。
        """
        return [
            {
                "字段名": "（列表收录）",
                "含义": "本表写入哪些自选股",
                "计算/取值规则": (
                    "仅用户自选中的 A 股；正式买点（是否买点=是）一律收录；"
                    "非买点须满足连阳硬筛：4日阳≥3 或 5日阳≥4（默认 yang_rule_a/b）。"
                    "排序：买点优先，再按得分降序。"
                ),
            },
            {
                "字段名": "股票代码",
                "含义": "A 股六位证券代码",
                "计算/取值规则": "自选股代码规范化为 6 位；港股/5 位码不进入本表。",
            },
            {
                "字段名": "股票名称",
                "含义": "证券简称",
                "计算/取值规则": "优先取策略返回名称，否则用自选股名称。",
            },
            {
                "字段名": "信号日",
                "含义": "指标与买点判定所依据的交易日",
                "计算/取值规则": "取策略 signal_date（或 date）；缺省为报告生成日 search_date。",
            },
            {
                "字段名": "收盘",
                "含义": "信号日收盘价",
                "计算/取值规则": "信号日 K 线收盘价，保留 2 位小数。",
            },
            {
                "字段名": "MA20",
                "含义": "收盘价简单移动平均（默认 20 日）",
                "计算/取值规则": (
                    "含信号日在内最近 ma_period（默认 20）根收盘价算术平均；"
                    "硬筛要求收盘 ≥ MA20（站上均线）。"
                ),
            },
            {
                "字段名": "4日阳",
                "含义": "近 4 个交易日阳线根数",
                "计算/取值规则": (
                    "含信号日向前共 4 日，阳线判定 close>open；"
                    "规则 A 默认要求 ≥3（yang_rule_a）。"
                ),
            },
            {
                "字段名": "5日阳",
                "含义": "近 5 个交易日阳线根数",
                "计算/取值规则": (
                    "含信号日向前共 5 日，阳线判定 close>open；"
                    "规则 B 默认要求 ≥4（yang_rule_b）。连阳硬筛为规则 A 或 B 任一通过。"
                ),
            },
            {
                "字段名": "10日阳",
                "含义": "近 10 个交易日阳线根数",
                "计算/取值规则": (
                    "中期阳线硬筛窗口之一；默认阈值 ≥6（use_yang_medium=true 时须与 15/20 日窗口一并满足）。"
                ),
            },
            {
                "字段名": "15日阳",
                "含义": "近 15 个交易日阳线根数",
                "计算/取值规则": "中期阳线硬筛窗口之一；默认阈值 ≥8。",
            },
            {
                "字段名": "20日阳",
                "含义": "近 20 个交易日阳线根数",
                "计算/取值规则": "中期阳线硬筛窗口之一；默认阈值 ≥10。",
            },
            {
                "字段名": "多头",
                "含义": "短中期均线是否多头排列",
                "计算/取值规则": (
                    "是=MA5>MA10>MA20；缺省 require_ma_bull=true 时作为硬筛条件；"
                    "空头排列另以风险标签提示。"
                ),
            },
            {
                "字段名": "量能倍数",
                "含义": "当日成交量相对近均量的倍数",
                "计算/取值规则": (
                    "当日成交量 / 过去 volume_lookback（默认 20）日均量（不含当日）；"
                    "硬筛默认要求 ≥ volume_multiple（缺省 3.0）；打分约在 4.0 倍附近拉满分项。"
                ),
            },
            {
                "字段名": "量比",
                "含义": "相对前日成交量之比（近似量比）",
                "计算/取值规则": (
                    "当日成交量 / 前一日成交量；默认不参与硬筛与加分"
                    "（use_volume_ratio=false）；开启后须 ≥ min_volume_ratio。"
                ),
            },
            {
                "字段名": "换手%",
                "含义": "信号日换手率（%）",
                "计算/取值规则": (
                    "取行情库当日 turnover_rate；缺省硬筛 ≥ min_turnover（默认 3%）；"
                    "积分按相对近 20 日中位甜区加分，过高则减分并可能打风险标签。"
                ),
            },
            {
                "字段名": "支撑",
                "含义": "信号日最近结构支撑价",
                "计算/取值规则": (
                    "与个股关键价位同口径（结构窗 KDE + 共振）；取 nearest_support；"
                    "用于结构硬闸与结构出场止损参考。"
                ),
            },
            {
                "字段名": "阻力",
                "含义": "信号日最近结构阻力价",
                "计算/取值规则": (
                    "同口径取 nearest_resistance；贴/超阻力或上行空间不足（缺省 <3%）可硬闸否决买点。"
                ),
            },
            {
                "字段名": "结构盈亏比",
                "含义": "结构上行空间 / 下行风险（无量纲 RR）",
                "计算/取值规则": (
                    "RR=上行/max(价−支撑, 现价×下限%, k×ATR)；偏低时仅软提示，不单独否决买点；"
                    "参与位置与 RR 得分分项。"
                ),
            },
            {
                "字段名": "风险",
                "含义": "结构/趋势/过热/换手等风险标签",
                "计算/取值规则": (
                    "多标签以「；」拼接，如：破位支撑、贴/超阻力、上行空间不足、悬空离支撑、"
                    "结构盈亏比偏低、近期涨幅偏大/过大、均线乖离偏大/过大、空头趋势、换手过高等。"
                    "其中破位/贴阻力/上行不足/悬空/涨幅过大/乖离过大等可构成硬闸否决正式买点。"
                ),
            },
            {
                "字段名": "得分",
                "含义": "URT 综合得分（上限 100）",
                "计算/取值规则": (
                    "分项合计后封顶 100（约）：MA20 趋势≤10；连阳天数≤20；K 线实体质量≤10；"
                    "量能倍数≤25；中期阳线≤5；均线多头 +0～8 / 空头 −8；筹码位置与 RR≤15；"
                    "换手 +8/−8；可选量比≤5；过热扣分约 −10～0。"
                    "正式买点另要求得分 ≥ min_score（默认 70）。满分≠已贴近最优买点。"
                ),
            },
            {
                "字段名": "是否买点",
                "含义": "当日是否发出 URT 正式买点",
                "计算/取值规则": (
                    "是=硬筛全部通过 ∧ 结构硬闸通过 ∧ 过热硬闸通过 ∧ 得分≥min_score；"
                    "硬筛缺省：站上 MA20 ∧ 连阳(4日≥3或5日≥4) ∧ 量能倍数≥阈值 ∧ 中期阳线 ∧ 多头 ∧ 换手≥3%。"
                    "否=未达正式买点，但仍可能因连阳条件被本表收录（观察补充，不可作回测入场）。"
                ),
            },
            {
                "字段名": "买点建议",
                "含义": "基于结构位的短线执行提示摘要",
                "计算/取值规则": (
                    "来自 trade_advice：动作 + summary 摘要（截断）；"
                    "正式买点可提示现价跟进或回踩承接；未达买点多为观察类文案。"
                ),
            },
        ]

    @staticmethod
    def _urt_report_excel_row(r: Dict[str, Any], *, report_date: str, code_to_name: Dict[str, str]) -> Dict[str, Any]:
        """对齐选股页列 + 正式买点/建议：中期阳线、多头、支撑/阻力、RR、风险等。"""
        code = str(r.get("code") or "").strip()
        if code.isdigit():
            code = code.zfill(6)
        name = (r.get("name") or "").strip() or code_to_name.get(code, "")
        signal_date = r.get("signal_date") or r.get("date") or report_date

        def _round(v: Any, n: int):
            if v is None or v == "":
                return ""
            try:
                return round(float(v), n)
            except (TypeError, ValueError):
                return ""

        def _int_or_blank(v: Any):
            if v is None or v == "":
                return ""
            try:
                return int(v)
            except (TypeError, ValueError):
                return ""

        ma_bull = r.get("ma_bull_ok")
        if ma_bull is True:
            ma_bull_txt = "是"
        elif ma_bull is False:
            ma_bull_txt = "否"
        else:
            ma_bull_txt = ""

        return {
            "股票代码": "\u2060" + str(code),
            "股票名称": name,
            "信号日": str(signal_date)[:10] if signal_date else "",
            "收盘": _round(r.get("close"), 2),
            "MA20": _round(r.get("ma20"), 2),
            "4日阳": _int_or_blank(r.get("yang_count_4")),
            "5日阳": _int_or_blank(r.get("yang_count_5")),
            "10日阳": _int_or_blank(r.get("yang_count_10")),
            "15日阳": _int_or_blank(r.get("yang_count_15")),
            "20日阳": _int_or_blank(r.get("yang_count_20")),
            "多头": ma_bull_txt,
            "量能倍数": _round(r.get("volume_multiple"), 2),
            "量比": _round(r.get("volume_ratio"), 2),
            "换手%": _round(r.get("turnover_rate"), 2),
            "支撑": _round(ReportService._urt_pick_support(r), 2),
            "阻力": _round(ReportService._urt_pick_resistance(r), 2),
            "结构盈亏比": _round(ReportService._urt_pick_structure_rr(r), 2),
            "风险": ReportService._urt_format_risk_tags(r),
            "得分": _round(r.get("score"), 1),
            "是否买点": "是" if bool(r.get("buy_signal")) else "否",
            "买点建议": ReportService._urt_format_trade_advice(r),
        }

    def _generate_urt_report_for_user(
        self, user_id: int, stock_codes: Optional[List[str]] = None
    ) -> ReportResult:
        """
        生成该用户自选股范围内的 URT 上升趋势策略报告（仅 A 股）。

        收录：
        1) 当日买点；
        2) 无买点但满足连阳硬筛（4日≥3阳 或 5日≥4阳）的自选股。
        列对齐选股页：代码/名称/信号日/收盘/MA20/4～20日阳/多头/量能倍数/量比/换手%/支撑/阻力/结构盈亏比/风险/得分/是否买点/买点建议。

        stock_codes: 推送任务可选子集；None 表示该用户全部自选股。
        """
        watchlist = self.get_user_watchlist(user_id, stock_codes)
        if not watchlist:
            return ReportResult(
                success=True,
                file_path=None,
                report_info=ReportInfo(
                    stock_count=0,
                    report_date=datetime.now().strftime("%Y-%m-%d"),
                    report_type="urt_daily",
                    file_size=0,
                    has_data=False,
                    missing_data_stocks=[],
                ),
                error_message="用户没有自选股",
            )

        def _norm_cn_code(raw: str) -> str:
            """A 股代码规范化：禁止把 5 位港股码 zfill(6) 抬成 A 股（00981→000981）。"""
            s = str(raw or "").replace("\u2060", "").strip()
            if not s or not s.isdigit():
                return s
            if len(s) == 5:
                return ""
            if len(s) > 6:
                return ""
            return s.zfill(6)

        stock_pool: List[str] = []
        code_to_name: Dict[str, str] = {}
        skipped_hk_like: List[str] = []
        for item in watchlist:
            market = (item.get("market") or "").strip().upper()
            raw_code = str(item.get("stock_code") or "").replace("\u2060", "").strip()
            # 港股 / 5 位数字码不得进入 A 股 URT 池（否则 00981 会被当成 000981）
            if market == "HK" or (raw_code.isdigit() and len(raw_code) == 5):
                if len(skipped_hk_like) < 20:
                    skipped_hk_like.append(raw_code)
                continue
            if market and market != "CN":
                continue
            code = _norm_cn_code(raw_code)
            if not code or not code.isdigit() or len(code) != 6:
                continue
            stock_pool.append(code)
            code_to_name[code] = item.get("stock_name") or ""

        if not stock_pool:
            return ReportResult(
                success=True,
                file_path=None,
                report_info=ReportInfo(
                    stock_count=0,
                    report_date=datetime.now().strftime("%Y-%m-%d"),
                    report_type="urt_daily",
                    file_size=0,
                    has_data=False,
                    missing_data_stocks=[],
                ),
                error_message="自选股中无 A 股标的",
            )

        logger.info(
            "URT 日报 user_id=%s A股自选池=%s 只 sample=%s skipped_hk_like=%s",
            user_id,
            len(stock_pool),
            stock_pool[:8],
            skipped_hk_like,
        )

        try:
            from backend_core.strategies.urt import URTFrontendInterface

            # 跳过选股硬筛，拿到自选全量可算明细（含非买点），再按推送规则过滤
            payload = URTFrontendInterface.screen(
                self.db,
                scope="watchlist",
                stock_codes=stock_pool,
                prefer_cache=False,
                force_realtime=True,
                skip_screening_filters=True,
            )
        except Exception as e:
            logger.exception("URT 自选股选股失败: %s", e)
            return ReportResult(
                success=False,
                file_path=None,
                report_info=None,
                error_message=str(e),
            )

        raw_results = payload.get("data") or [] if isinstance(payload, dict) else []
        report_date = (
            str(payload.get("search_date") or "")[:10]
            if isinstance(payload, dict) and payload.get("search_date")
            else datetime.now().strftime("%Y-%m-%d")
        )

        allow = set(stock_pool)
        in_pool: List[Dict[str, Any]] = []
        leaked = 0
        leaked_codes: List[str] = []
        for r in raw_results:
            if not isinstance(r, dict):
                continue
            code = _norm_cn_code(str(r.get("code") or "").replace("\u2060", "").strip())
            if code not in allow:
                leaked += 1
                if len(leaked_codes) < 20:
                    leaked_codes.append(code)
                continue
            in_pool.append(r)
        logger.info(
            "URT 日报 user_id=%s screen返回=%s 池内=%s 剔除非自选=%s leaked_sample=%s",
            user_id,
            len(raw_results),
            len(in_pool),
            leaked,
            leaked_codes,
        )
        if leaked:
            logger.warning(
                "URT 日报 user_id=%s 过滤掉非自选股结果 %s 条（应仅含自选池） codes=%s",
                user_id,
                leaked,
                leaked_codes,
            )

        results = [r for r in in_pool if self._urt_row_meets_report_filter(r)]
        results.sort(
            key=lambda x: (
                1 if bool(x.get("buy_signal")) else 0,
                float(x.get("score") or 0),
            ),
            reverse=True,
        )

        if not results:
            return ReportResult(
                success=True,
                file_path=None,
                report_info=ReportInfo(
                    stock_count=0,
                    report_date=report_date,
                    report_type="urt_daily",
                    file_size=0,
                    has_data=False,
                    missing_data_stocks=[],
                ),
                error_message="当日自选股范围内无 URT 买点，且无满足连阳条件（4日≥3阳或5日≥4阳）的标的",
            )

        rows = [
            self._urt_report_excel_row(r, report_date=report_date, code_to_name=code_to_name)
            for r in results
        ]

        # 文件名带生成时刻，避免与历史同日报告混淆（微信下载出现 urt_9_20260731(1).xlsx）
        gen_ts = datetime.now().strftime("%H%M%S")
        filename = f"urt_{user_id}_{report_date.replace('-', '')}_{gen_ts}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        df = pd.DataFrame(rows)
        if "股票代码" in df.columns:
            df["股票代码"] = df["股票代码"].astype(str)
        legend_df = pd.DataFrame(self._urt_report_field_legend_rows())
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="URT策略信号列表")
            legend_df.to_excel(writer, index=False, sheet_name="字段说明")
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(filepath)
            ws = wb["URT策略信号列表"] if "URT策略信号列表" in wb.sheetnames else wb[wb.sheetnames[0]]
            title_to_col = {}
            for cell in ws[1]:
                if cell.value is not None:
                    title_to_col[str(cell.value).strip()] = cell.column
            desired = {
                "股票代码": 12,
                "股票名称": 14,
                "信号日": 12,
                "收盘": 9,
                "MA20": 9,
                "4日阳": 8,
                "5日阳": 8,
                "10日阳": 8,
                "15日阳": 8,
                "20日阳": 8,
                "多头": 6,
                "量能倍数": 10,
                "量比": 8,
                "换手%": 8,
                "支撑": 9,
                "阻力": 9,
                "结构盈亏比": 10,
                "风险": 28,
                "得分": 8,
                "是否买点": 10,
                "买点建议": 36,
            }
            for title, width in desired.items():
                col_idx = title_to_col.get(title)
                if col_idx:
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
            if "字段说明" in wb.sheetnames:
                ws_legend = wb["字段说明"]
                legend_widths = {"字段名": 14, "含义": 28, "计算/取值规则": 72}
                legend_title_to_col = {}
                for cell in ws_legend[1]:
                    if cell.value is not None:
                        legend_title_to_col[str(cell.value).strip()] = cell.column
                for title, width in legend_widths.items():
                    col_idx = legend_title_to_col.get(title)
                    if col_idx:
                        ws_legend.column_dimensions[get_column_letter(col_idx)].width = width
            wb.save(filepath)
        except Exception as e:
            logger.warning("URT 报告 Excel 列宽样式应用失败: %s", e)

        buy_n = sum(1 for r in results if bool(r.get("buy_signal")))
        file_size = os.path.getsize(filepath)
        logger.info(
            "生成 URT 自选股报告成功: %s, 总行数=%s, 买点=%s, 阳线补充=%s",
            filepath,
            len(rows),
            buy_n,
            len(rows) - buy_n,
        )
        return ReportResult(
            success=True,
            file_path=filepath,
            report_info=ReportInfo(
                stock_count=len(rows),
                report_date=report_date,
                report_type="urt_daily",
                file_size=file_size,
                has_data=True,
                missing_data_stocks=[],
            ),
            error_message=None,
        )

    def get_report_info(self, report_path: str) -> Optional[ReportInfo]:
        """
        获取报告信息
        
        Args:
            report_path: 报告文件路径
            
        Returns:
            ReportInfo: 报告信息，如果文件不存在返回None
        """
        try:
            if not os.path.exists(report_path):
                logger.error(f"报告文件不存在: {report_path}")
                return None
            
            # 获取文件大小
            file_size = os.path.getsize(report_path)
            
            # 从文件名解析报告类型和日期
            filename = os.path.basename(report_path)
            if 'volume_aberration' in filename:
                report_type = 'volume_aberration'
            elif 'gms_' in filename:
                report_type = 'gms_daily'
            elif 'urt_' in filename:
                report_type = 'urt_daily'
            elif 'summary' in filename:
                report_type = 'summary'
            else:
                report_type = 'detailed'
            report_date = datetime.now().strftime("%Y-%m-%d")
            
            # 读取文件获取股票数量
            stock_count = 0
            has_data = False
            
            if report_path.endswith('.csv'):
                df = pd.read_csv(report_path, encoding='utf-8-sig')
                stock_count = len(df)
                has_data = stock_count > 0
            elif report_path.endswith('.xlsx'):
                if report_type == 'volume_aberration':
                    try:
                        df_cn = pd.read_excel(report_path, sheet_name='A股放量榜')
                        stock_count += len(df_cn)
                    except Exception:
                        pass
                    try:
                        df_hk = pd.read_excel(report_path, sheet_name='港股放量榜')
                        stock_count += len(df_hk)
                    except Exception:
                        pass
                    has_data = stock_count > 0
                elif report_type == 'gms_daily':
                    # GMS 推送报告默认写入该 sheet；兼容历史/异常文件时回退到首个 sheet
                    try:
                        df = pd.read_excel(report_path, sheet_name='GMS策略信号列表')
                    except Exception:
                        xls = pd.ExcelFile(report_path)
                        first_sheet = xls.sheet_names[0] if xls.sheet_names else 0
                        df = pd.read_excel(report_path, sheet_name=first_sheet)
                    stock_count = len(df)
                    has_data = stock_count > 0
                elif report_type == 'urt_daily':
                    try:
                        df = pd.read_excel(report_path, sheet_name='URT策略信号列表')
                    except Exception:
                        xls = pd.ExcelFile(report_path)
                        first_sheet = xls.sheet_names[0] if xls.sheet_names else 0
                        df = pd.read_excel(report_path, sheet_name=first_sheet)
                    stock_count = len(df)
                    has_data = stock_count > 0
                else:
                    df = pd.read_excel(report_path, sheet_name='股票汇总')
                    stock_count = len(df)
                    has_data = stock_count > 0
            
            return ReportInfo(
                stock_count=stock_count,
                report_date=report_date,
                report_type=report_type,
                file_size=file_size,
                has_data=has_data,
                missing_data_stocks=[]  # 从文件中无法获取，返回空列表
            )
            
        except Exception as e:
            logger.error(f"获取报告信息失败: {str(e)}")
            return None
