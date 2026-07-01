from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).with_name("GMS策略功能模块完成列表.xlsx").resolve()


THEME = {
    "header": "1F4E78",
    "subheader": "D9EAF7",
    "done": "D9EAD3",
    "partial": "FFF2CC",
    "doing": "DDEBF7",
    "planned": "EADCF8",
    "risk": "F4CCCC",
    "text": "1F2937",
}


overview_rows = [
    ["项目名称", "GMS（均值引力与动量突变策略）", "基于 mean_frequency_resonance_indicators 的双模块阶梯式评分选股策略"],
    ["整理日期", datetime.now().strftime("%Y-%m-%d"), "由本地文档、代码目录和 GMS 相关 API/页面整理"],
    ["总体完成度", "约 95%", "核心算法、数据、选股、回测、管理端、前端展示已具备生产使用条件"],
    ["核心能力", "左侧均值吸附 + 右侧动量引爆", "同时支持左侧买点、右侧买点、卖点/过热信号、得分明细"],
    ["覆盖市场", "A股 / 港股 / ETF / 自选股 / GMS观察股 / 行业板块 / 概念板块", "以当前前端筛选范围和后端股票池支持为准"],
    ["主要待完善", "性能监控、告警、备份、部分公共参数持久化体验", "不影响主流程使用，但建议纳入后续迭代"],
]


module_rows = [
    ["一级模块", "二级模块", "功能项", "完成状态", "完成度", "主要说明", "关键文件/页面", "备注"],
    ["策略核心引擎", "数据模型层", "GMSIndicators、GMSSignal、异常类型、双模块评分字段", "已完成", "100%", "策略输入、输出、信号结构完整", "backend_core/strategies/gms/models.py", ""],
    ["策略核心引擎", "配置管理", "默认参数、配置文件加载、深度合并、左/右侧买点参数、评分参数", "已完成", "100%", "支持 default / gms_penalty 等参数版本管理", "backend_core/strategies/gms/config.py; gms_config.json", ""],
    ["策略核心引擎", "数据加载器", "读取 mean_frequency_resonance_indicators，支持 CN/HK、最近可用日、多日数据", "已完成", "100%", "为站稳 3 日、批量筛选、单股追溯提供数据基础", "backend_core/strategies/gms/data_loader.py", ""],
    ["策略核心引擎", "指标计算器", "均值收敛态、动量溢出态、S/A 等级、全速/分批等级", "已完成", "100%", "双模块阶梯式评分算法已实现", "backend_core/strategies/gms/indicators_calculator.py; backend_core/strategies/gms/scoring/", ""],
    ["策略核心引擎", "信号检测器", "左侧买点、右侧买点、卖点、等级优先与兜底逻辑", "已完成", "100%", "可输出买卖信号、原因和条件明细", "backend_core/strategies/gms/signal_detector.py", ""],
    ["策略核心引擎", "策略引擎", "完整选股流程、批量处理、市场筛选、最低分、结果限制", "已完成", "100%", "串联数据加载、评分、信号检测、结果排序", "backend_core/strategies/gms/strategy_engine.py", ""],
    ["策略核心引擎", "前端接口封装", "统一 API 调用入口、股票池获取、结果格式化", "已完成", "100%", "为选股页、详情页、回测扫描提供统一封装", "backend_core/strategies/gms/frontend_interface.py", ""],
    ["数据采集与指标", "历史行情采集", "A股、港股历史行情采集并计算 GMS 所需基础指标", "已完成", "95%", "A股 AkShare/TuShare、港股 AkShare 均已覆盖", "backend_core/data_collectors/akshare/*historical*.py; backend_core/data_collectors/tushare/historical.py", ""],
    ["数据采集与指标", "指标计算入库", "均值频率共振指标、宏观位移、偏离率、突变率、涨跌天数、量能指标", "已完成", "95%", "历史采集、ETF 采集、观察股采集均有 GMS 指标计算入口", "backend_core/data_collectors/akshare/historical_collector.py; etf_collector.py", ""],
    ["数据采集与指标", "MA60_D 来源同步", "ma60_d 回填和 lookup，支持 GMS 打分机制增强", "已完成", "100%", "已提供迁移与同步脚本", "backend_core/strategies/gms/ma60_source.py; migrations/sync_mfr_ma60_d_from_ma_indicators.py", ""],
    ["数据库/迁移", "指标主表", "mean_frequency_resonance_indicators 主表与字段", "已完成", "95%", "支持 CN/HK 市场和 d1/d20/ratio 等核心字段", "migrations/*mean_frequency*; backend_api/models.py", ""],
    ["数据库/迁移", "策略配置版本", "gms_strategy_configs、config_id 绑定、默认版本/减分版", "已完成", "100%", "支持参数多版本、克隆、默认版本、停用", "migrations/add_gms_strategy_configs.py; consolidate_gms_canonical_configs.py", ""],
    ["数据库/迁移", "打分机制字段", "scoring.mechanism、减分规则、版本绑定补全", "已完成", "100%", "支持标准版与减分版机制", "migrations/add_gms_scoring_mechanism.py; backend_core/strategies/gms/scoring/", ""],
    ["数据库/迁移", "交易观察/正式交易表", "gms_trade_observe_stocks、history、gms_formal_trades", "已完成", "100%", "支持网站交易观察、移除归档、正式交易记录", "migrations/add_gms_trade_observe_*.py; add_gms_formal_trades.py", ""],
    ["选股 API", "GMS 筛选接口", "GMS 策略筛选、单股、全部 A 股、港股、ETF、自选、板块等范围", "已完成", "95%", "前端选股页调用主入口，支持分页和 trace_only", "backend_api/stock/gms_frontend_routes.py; frontend/js/screening.js", ""],
    ["选股 API", "选股结果查询", "基于 gms_signal_trace 的公开/管理端选股结果", "已完成", "100%", "管理端与前端查询同源逻辑", "backend_api/services/gms_signal_trace_selection.py; backend_api/admin/gms_admin_routes.py", ""],
    ["选股 API", "参数读取与保存", "默认配置、策略参数版本、比较、克隆、更新、默认设置", "已完成", "95%", "管理端配置已持久化；前端仍保留 localStorage 兜底", "backend_api/admin/gms_admin_routes.py; frontend/js/screening.js", "旧文档中“服务端持久化”已基本补齐"],
    ["前端页面", "选股页面", "GMS 策略标签页、说明、参数、范围选择、结果表格", "已完成", "95%", "已支持多数据来源、分页、得分明细、操作按钮", "frontend/screening.html; frontend/js/screening.js; frontend/css/screening.css", ""],
    ["前端页面", "得分明细", "双模块评分、扣分、维度细项、展开/折叠、Excel 注释文本", "已完成", "100%", "选股页与追溯页共用构建逻辑", "frontend/js/gms_score_detail.js; frontend/js/screening.js", ""],
    ["前端页面", "导出功能", "GMS 筛选结果 CSV / Excel 导出", "已完成", "100%", "Excel 导出包含数据行和得分明细行", "frontend/js/screening.js", ""],
    ["前端页面", "个股详情 GMS 卡片", "股票详情页加载 GMS 指标与信号展示", "已完成", "95%", "A股/港股详情页均有 GMS 指标入口", "frontend/js/stock.js; frontend/js/stock_hk.js; frontend/css/stock.css", ""],
    ["信号追溯", "追溯页面", "单股 GMS 历史信号、指标明细、策略版本选择、重算", "已完成", "95%", "支持按配置版本重算历史交易日信号", "frontend/stock_gms_trace.html; frontend/js/stock_gms_trace.js; backend_api/stock/gms_trace_routes.py", ""],
    ["交易观察", "观察股列表", "从选股结果加入交易观察、列表、代码集合、历史归档", "已完成", "100%", "加入观察时同步保障 GMS 策略观察股", "backend_api/gms_trade_observe_routes.py; frontend/js/screening.js", ""],
    ["正式交易", "正式交易记录", "从观察股/选股结果进入正式交易记录管理", "已完成", "95%", "支持记录状态、代码、市场、买入信息等", "backend_api/gms_formal_trade_routes.py; frontend/js/screening.js", ""],
    ["管理端", "GMS 策略版本", "版本列表、启用、绑定参数版本、观察股管理、批量导入/删除", "已完成", "100%", "管理端侧栏 GMS策略版本", "admin/src/views/GmsWatchlistView.vue; backend_api/admin/gms_admin_routes.py", ""],
    ["管理端", "GMS 回测管理", "任务创建、列表、状态、日志、取消、重跑、删除", "已完成", "95%", "支持单股、自选、GMS观察股等回测范围", "admin/src/views/GMSManagementView.vue; backend_core/strategies/gms/backtest_*.py", ""],
    ["管理端", "回测报告", "报告列表、详情、下载、删除", "已完成", "95%", "支持报告持久化和历史任务迁移", "backend_api/admin/gms_admin_routes.py; backend_core/strategies/gms/backtest_storage.py", ""],
    ["管理端", "系统状态", "运行任务、报告统计、健康检查", "已完成", "90%", "提供基础监控状态，仍可继续增强告警", "backend_api/admin/gms_admin_routes.py", ""],
    ["推送/报告", "GMS 每日信号推送", "推送配置支持 gms_daily 报告类型", "已完成", "90%", "配置页面已有 GMS 报告类型", "admin/src/views/PushConfigView.vue; backend_core/scheduler/push_scheduler.py", ""],
    ["部署运维", "开发环境", "本地开发、数据库结构、API、前端可访问", "已完成", "95%", "以现有运行脚本和文档为准", "start_backend_api.py; start_frontend.py; 日常运维.md", ""],
    ["部署运维", "生产监控/告警/备份", "性能监控、错误报警、数据备份策略", "部分完成", "70%", "基础部署具备，专项监控与告警仍建议完善", "日常运维.md", "后续重点"],
]


todo_rows = [
    ["类别", "优化项", "当前状态", "优先级", "建议动作", "关联模块"],
    ["参数管理", "前端公共参数持久化体验统一", "部分完成", "高", "减少 localStorage 兜底依赖，统一使用服务端策略配置版本", "frontend/js/screening.js; backend_api/admin/gms_admin_routes.py"],
    ["性能优化", "大批量全 A 股筛选性能测试", "进行中/待验证", "高", "沉淀基准耗时、分页策略、超时配置和慢 SQL 定位", "backend_core/strategies/gms; backend_api/stock/gms_frontend_routes.py"],
    ["性能优化", "数据库查询优化和索引复核", "计划中", "高", "围绕 gms_signal_trace、mean_frequency_resonance_indicators、配置版本字段检查索引", "migrations; database"],
    ["缓存机制", "高频筛选结果缓存", "计划中", "中", "对相同日期、范围、config_id 的结果做缓存/预计算", "scheduled_precompute.py; backtest_storage.py"],
    ["回测增强", "策略效果统计维度扩展", "部分完成", "中", "增加胜率、盈亏比、最大回撤、持有周期分布等报表", "backtest_runner.py; GMSManagementView.vue"],
    ["风险提示", "信号风险提示增强", "计划中", "中", "结合过热、量能不足、趋势破坏等规则输出风险标签", "signal_detector.py; gms_score_detail.js"],
    ["监控告警", "策略执行监控和错误告警", "部分完成", "中", "接入定时任务、失败率、耗时、异常告警", "scheduler; 日常运维.md"],
    ["日志审计", "用户行为与操作审计", "计划中", "低", "记录参数修改、回测创建、观察股操作等关键动作", "backend_api/admin/gms_admin_routes.py"],
    ["文档维护", "更新旧完成清单中 PVFRS/GMS 命名混用内容", "计划中", "低", "把 PVFRS 历史命名统一标注为 GMS/PVFRS 指标来源", ".kiro/specs/gms-strategy/work-completion-list.md"],
]


file_rows = [
    ["分类", "路径", "用途", "状态"],
    ["核心引擎", "backend_core/strategies/gms/models.py", "GMS 数据模型、信号模型", "已完成"],
    ["核心引擎", "backend_core/strategies/gms/config.py", "配置管理、参数版本默认值、配置解析", "已完成"],
    ["核心引擎", "backend_core/strategies/gms/data_loader.py", "指标数据加载、市场和交易日兜底", "已完成"],
    ["核心引擎", "backend_core/strategies/gms/indicators_calculator.py", "双模块评分计算", "已完成"],
    ["核心引擎", "backend_core/strategies/gms/signal_detector.py", "买卖点信号检测", "已完成"],
    ["核心引擎", "backend_core/strategies/gms/strategy_engine.py", "策略选股主流程", "已完成"],
    ["核心引擎", "backend_core/strategies/gms/frontend_interface.py", "前端选股接口封装", "已完成"],
    ["核心引擎", "backend_core/strategies/gms/scoring/", "打分机制、扣分规则、机制元数据", "已完成"],
    ["回测", "backend_core/strategies/gms/backtest_runner.py", "GMS 回测执行", "已完成"],
    ["回测", "backend_core/strategies/gms/backtest_worker.py", "回测异步任务执行", "已完成"],
    ["回测", "backend_core/strategies/gms/backtest_storage.py", "回测任务/报告存储", "已完成"],
    ["预计算", "backend_core/strategies/gms/scheduled_precompute.py", "定时预计算", "已完成"],
    ["API", "backend_api/admin/gms_admin_routes.py", "管理端 GMS 回测、配置、版本、观察股 API", "已完成"],
    ["API", "backend_api/stock/gms_frontend_routes.py", "前端 GMS 选股/选股结果 API", "已完成"],
    ["API", "backend_api/stock/gms_trace_routes.py", "GMS 信号追溯 API", "已完成"],
    ["API", "backend_api/gms_trade_observe_routes.py", "GMS 交易观察 API", "已完成"],
    ["API", "backend_api/gms_formal_trade_routes.py", "GMS 正式交易 API", "已完成"],
    ["服务", "backend_api/services/gms_signal_trace_selection.py", "基于信号追溯表查询选股结果", "已完成"],
    ["服务", "backend_api/services/gms_strategy_watchlist.py", "GMS 策略观察股保障/同步", "已完成"],
    ["前端", "frontend/screening.html", "GMS 筛选页结构", "已完成"],
    ["前端", "frontend/js/screening.js", "GMS 筛选交互、导出、分页、操作", "已完成"],
    ["前端", "frontend/js/gms_score_detail.js", "GMS 得分明细组件", "已完成"],
    ["前端", "frontend/stock_gms_trace.html", "GMS 信号追溯页面", "已完成"],
    ["前端", "frontend/js/stock_gms_trace.js", "GMS 信号追溯交互", "已完成"],
    ["管理端", "admin/src/views/GmsWatchlistView.vue", "GMS 策略版本与观察股管理", "已完成"],
    ["管理端", "admin/src/views/GMSManagementView.vue", "GMS 回测管理中心", "已完成"],
    ["管理端", "admin/src/utils/gmsScreeningFormat.ts", "管理端/前端筛选格式对齐", "已完成"],
    ["迁移", "migrations/add_gms_strategy_configs.py", "GMS 策略参数多版本", "已完成"],
    ["迁移", "migrations/add_gms_scoring_mechanism.py", "GMS 打分机制字段", "已完成"],
    ["迁移", "migrations/add_gms_trade_observe_stocks.py", "GMS 交易观察表", "已完成"],
    ["迁移", "migrations/add_gms_trade_observe_history.py", "GMS 观察股移除归档", "已完成"],
    ["迁移", "migrations/add_gms_formal_trades.py", "GMS 正式交易表", "已完成"],
]


api_rows = [
    ["分类", "接口/页面", "方法/类型", "功能说明", "状态"],
    ["管理端 API", "/api/admin/gms/system/status", "GET", "GMS 系统状态、运行任务和报告统计", "已完成"],
    ["管理端 API", "/api/admin/gms/backtests", "GET/POST", "回测任务列表与创建", "已完成"],
    ["管理端 API", "/api/admin/gms/backtests/{task_id}", "GET/DELETE", "回测任务详情和删除", "已完成"],
    ["管理端 API", "/api/admin/gms/backtests/{task_id}/logs", "GET", "回测日志", "已完成"],
    ["管理端 API", "/api/admin/gms/backtests/{task_id}/cancel", "POST", "取消回测任务", "已完成"],
    ["管理端 API", "/api/admin/gms/backtests/{task_id}/rerun", "POST", "重新执行回测任务", "已完成"],
    ["管理端 API", "/api/admin/gms/reports", "GET", "回测报告列表", "已完成"],
    ["管理端 API", "/api/admin/gms/reports/{report_id}/download", "GET", "下载回测报告", "已完成"],
    ["管理端 API", "/api/admin/gms/config", "GET/PUT", "兼容旧版默认配置读取/保存", "已完成"],
    ["管理端 API", "/api/admin/gms/strategy-configs", "GET/POST", "策略参数版本列表/创建", "已完成"],
    ["管理端 API", "/api/admin/gms/strategy-configs/{config_id}", "GET/PUT", "策略参数版本详情/更新", "已完成"],
    ["管理端 API", "/api/admin/gms/strategy-configs/{config_id}/clone", "POST", "克隆策略参数版本", "已完成"],
    ["管理端 API", "/api/admin/gms/strategy-versions", "GET/POST", "GMS 策略版本列表/创建", "已完成"],
    ["管理端 API", "/api/admin/gms/strategy-versions/{version_id}", "GET/PUT/DELETE", "策略版本详情/更新/删除", "已完成"],
    ["管理端 API", "/api/admin/gms/strategy-version-stocks", "GET/POST", "版本观察股列表/新增", "已完成"],
    ["管理端 API", "/api/admin/gms/strategy-version-stocks/batch-import", "POST", "批量导入观察股", "已完成"],
    ["前端 API", "/api/stock/gms-trade-observe/list", "GET", "用户 GMS 交易观察列表", "已完成"],
    ["前端 API", "/api/stock/gms-trade-observe/add", "POST", "加入 GMS 交易观察", "已完成"],
    ["前端 API", "/api/stock/gms-trade-observe/history", "GET", "观察股移除历史", "已完成"],
    ["前端页面", "frontend/screening.html", "页面", "GMS 选股、交易观察、正式交易入口", "已完成"],
    ["前端页面", "frontend/stock_gms_trace.html", "页面", "GMS 单股信号追溯", "已完成"],
    ["管理端页面", "admin/src/views/GmsWatchlistView.vue", "页面", "GMS 策略版本与观察股维护", "已完成"],
    ["管理端页面", "admin/src/views/GMSManagementView.vue", "页面", "GMS 回测管理中心", "已完成"],
]


def style_sheet(ws, freeze=True):
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Microsoft YaHei", size=10, color=THEME["text"])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=THEME["header"])
        cell.font = Font(name="Microsoft YaHei", size=10, color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.auto_filter.ref = ws.dimensions


def status_fill(status):
    if "已完成" in status:
        return THEME["done"]
    if "部分" in status:
        return THEME["partial"]
    if "进行" in status or "待验证" in status:
        return THEME["doing"]
    if "计划" in status:
        return THEME["planned"]
    return "FFFFFF"


def write_table(wb, title, rows, widths=None, status_col=None):
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    if status_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col)
            cell.fill = PatternFill("solid", fgColor=status_fill(str(cell.value or "")))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    else:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 28 if row == 1 else 42
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)

    write_table(
        wb,
        "概览",
        [["项目", "内容", "说明"], *overview_rows],
        widths=[18, 42, 78],
    )
    write_table(
        wb,
        "功能模块完成列表",
        module_rows,
        widths=[18, 22, 36, 14, 12, 48, 58, 24],
        status_col=4,
    )
    write_table(
        wb,
        "待优化事项",
        todo_rows,
        widths=[16, 34, 18, 12, 52, 46],
        status_col=3,
    )
    write_table(
        wb,
        "代码文件映射",
        file_rows,
        widths=[16, 58, 52, 14],
        status_col=4,
    )
    write_table(
        wb,
        "API与页面清单",
        api_rows,
        widths=[16, 52, 16, 52, 14],
        status_col=5,
    )

    # Add a small summary block to the overview sheet.
    ws = wb["概览"]
    start = ws.max_row + 3
    summary = [
        ["完成度分类", "模块数量"],
        ["已完成", sum(1 for r in module_rows[1:] if r[3] == "已完成")],
        ["部分完成", sum(1 for r in module_rows[1:] if r[3] == "部分完成")],
        ["合计", len(module_rows) - 1],
    ]
    for row in summary:
        ws.append(row)
    for row in range(start, ws.max_row + 1):
        for col in range(1, 3):
            c = ws.cell(row, col)
            c.border = Border(left=Side(style="thin", color="D0D7DE"), right=Side(style="thin", color="D0D7DE"), top=Side(style="thin", color="D0D7DE"), bottom=Side(style="thin", color="D0D7DE"))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.font = Font(name="Microsoft YaHei", size=10, bold=(row == start))
            if row == start:
                c.fill = PatternFill("solid", fgColor=THEME["subheader"])

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    build()
