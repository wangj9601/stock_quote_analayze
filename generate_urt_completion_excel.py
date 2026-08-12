"""生成 URT 上升趋势策略功能模块列表（业务视角，不含代码路径）。"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "exported_docs" / "URT上升趋势策略功能模块列表.xlsx"

THEME = {
    "header": "1F4E78",
    "subheader": "D9EAF7",
    "done": "D9EAD3",
    "partial": "FFF2CC",
    "doing": "DDEBF7",
    "planned": "EADCF8",
    "todo": "FCE4D6",
    "text": "1F2937",
}

overview_rows = [
    ["项目名称", "URT（上升趋势策略 / Upward Right-side Trend）", "右侧趋势选股：站上 MA20 + 连阳 + 量能确认"],
    ["整理日期", datetime.now().strftime("%Y-%m-%d"), "以代码与 docs/strategies/urt/URT_GMS功能对比与技术方案.md 为准"],
    ["总体完成度", "约 92%", "港股选股/预计算、交易观察、日报、选股版本 UI、risk_exit 已落地"],
    ["策略定位", "右侧上升趋势确认", "硬筛通过后再按百分制得分过滤；不做左侧吸附"],
    ["覆盖范围", "A股 / 港股全市场 / 自选 / 行业概念 / 单股(仅A股)", "ETF、GMS式版本观察股池未覆盖；日报仅自选A股"],
    ["数据主源", "日 K 现算", "historical_quotes / historical_quotes_hk；可选读 urt_signal_trace"],
    ["主要待完善", "ETF scope、版本观察股池；港股单股/日报/回测待产品确认", "详见对比文 §4.2"],
]

# 一级模块 | 二级模块 | 功能说明 | 完成状态 | 完成度 | 备注
module_rows = [
    ["一级模块", "二级模块", "功能说明", "完成状态", "完成度", "备注"],
    # ── 一、策略算法 ──
    ["策略算法", "硬筛规则", "收盘站上 MA20、连阳（4日≥3阳或5日≥4阳）、量能达阈值倍数", "已完成", "100%", ""],
    ["策略算法", "百分制得分", "MA20/连阳/量能主分项；可选换手、量比加分", "已完成", "100%", "默认满分结构约 90 分"],
    ["策略算法", "买点信号判定", "硬筛通过且得分≥最低分后输出买点", "已完成", "100%", "单股模式可跳过滤输出明细"],
    ["策略算法", "信号原因明细", "得分拆解、买点逻辑说明可追溯展示", "已完成", "100%", ""],
    ["策略算法", "离场纪律函数", "价格止损、连跌离场、高点回撤止盈等规则实现", "已完成", "100%", "回测 exit_mode=risk_exit 可调用；默认 hit_rate"],
    # ── 二、数据基础 ──
    ["数据基础", "日 K 行情依赖", "A 股 historical_quotes；港股 historical_quotes_hk", "已完成", "100%", "与 GMS 指标链路不同，属设计差异"],
    ["数据基础", "候选股票池", "从股票基本信息过滤 ST、采集开关、代码位数等", "已完成", "100%", "ST 为候选池硬剔除，无选股勾选开关"],
    ["数据基础", "板块前缀过滤", "创业板/科创/沪深主板/中小/北证等代码段过滤", "已完成", "100%", "可多选并集"],
    # ── 三、选股能力 ──
    ["选股能力", "全 A 股筛选", "全部 A 股范围扫描上升趋势信号", "已完成", "95%", "可设扫描上限与超时"],
    ["选股能力", "自选股筛选", "按当前登录用户自选股池筛选", "已完成", "100%", "空自选=空结果"],
    ["选股能力", "行业板块选股", "按行业板块成分股多选筛选", "已完成", "100%", ""],
    ["选股能力", "概念板块选股", "按概念板块成分股多选筛选", "已完成", "100%", ""],
    ["选股能力", "A 股板块分段", "主板/创业板/中小板/科创板/北证等分段", "已完成", "100%", ""],
    ["选股能力", "单股查询模式", "输入代码实时计算信号明细（可含未过筛结果）", "部分完成", "80%", "暂仅 6 位 A 股"],
    ["选股能力", "缓存优先选股", "无临时参数覆盖时优先读预计算信号表", "已完成", "95%", ""],
    ["选股能力", "得分明细展示", "选股页展开查看分项得分与买点逻辑", "已完成", "100%", ""],
    ["选股能力", "结果导出", "筛选结果导出 CSV", "已完成", "95%", ""],
    ["选股能力", "策略参数版本切换", "选股时可指定不同参数版本", "已完成", "100%", "前台 urt-config_id + config_id"],
    ["选股能力", "港股范围", "港股全市场选股 scope=hk", "已完成", "95%", ""],
    ["选股能力", "ETF 范围", "独立 ETF 市场范围筛选", "未实现", "0%", "GMS 有对等能力"],
    ["选股能力", "观察股池范围", "独立策略版本观察股池作为选股范围", "未实现", "0%", "有用户侧交易观察，非版本池"],
    # ── 四、信号追溯与预计算 ──
    ["信号追溯", "历史信号查询", "单股按日期查看历史 URT 信号", "已完成", "95%", ""],
    ["信号追溯", "多版本切换", "历史页可切换参数版本查看", "已完成", "100%", ""],
    ["信号追溯", "信号重算", "指定区间异步强制重算，支持进度查询", "已完成", "100%", ""],
    ["信号追溯", "计算明细页", "单股信号计算明细独立页面", "已完成", "100%", ""],
    ["信号追溯", "追溯页单股回测", "在历史页发起单股回测", "已完成", "100%", "POST /api/stock/urt-backtest"],
    ["信号预计算", "定时全 A 预计算", "收盘后对全 A 股写入信号缓存", "已完成", "95%", "默认约 16:45"],
    ["信号预计算", "手动触发预计算", "管理端可手动跑预计算", "已完成", "100%", "market=CN|HK"],
    ["信号预计算", "港股预计算", "港股日终信号预计算", "已完成", "95%", "默认约 17:20"],
    # ── 五、回测验证 ──
    ["回测验证", "目标命中率回测", "次日开盘入场，观察期内达目标涨幅判定命中", "已完成", "95%", "当前默认模式；仅 A 股"],
    ["回测验证", "多范围回测", "支持全市场/自选等范围批量回测", "已完成", "90%", ""],
    ["回测验证", "任务管理", "创建、查看、取消、重跑、删除回测任务", "已完成", "95%", ""],
    ["回测验证", "报告与下载", "报告列表、详情、CSV / xlsx 导出", "已完成", "100%", ""],
    ["回测验证", "止损纪律回测模式", "持仓期按止损/连跌/回撤止盈出场", "已完成", "95%", "exit_mode=risk_exit"],
    # ── 六、策略配置 ──
    ["策略配置", "参数多版本管理", "策略参数版本创建、编辑、默认版本", "已完成", "95%", "urt_strategy_configs"],
    ["策略配置", "预计算开关", "按版本控制是否参与预计算", "已完成", "100%", ""],
    ["策略配置", "管理端试算预览", "配置页可试算选股预览", "已完成", "100%", ""],
    ["策略配置", "观察股版本分组", "按版本维护观察股池并绑定参数", "未实现", "0%", "GMS 有对等能力"],
    # ── 七、交易执行链 ──
    ["交易执行", "交易观察股", "从选股结果加入观察、列表管理、移除归档", "已完成", "95%", "urt_trade_observe*"],
    ["交易执行", "正式交易记录", "从观察进入正式交易并记录仓位状态", "已完成", "95%", "urt_formal_trade*"],
    ["交易执行", "选股页三子面板", "策略信号 / 交易观察 / 正式交易一体", "已完成", "100%", ""],
    # ── 八、用户端界面 ──
    ["用户端", "选股页上升趋势标签", "范围、参数、结果表、历史/明细入口", "已完成", "95%", "含刷新筛选"],
    ["用户端", "权限控制", "Tab/刷新/导出权限码控制可见性", "已完成", "100%", ""],
    ["用户端", "信号追溯独立页", "单股历史信号与重算", "已完成", "95%", ""],
    ["用户端", "个股详情 URT 卡片", "个股详情页展示 URT 指标与信号", "已完成", "100%", "stock.js + urt-score-detail"],
    # ── 九、管理端运营 ──
    ["管理端", "URT 管理中心", "策略参数、回测管理、报告与分析三 Tab", "已完成", "95%", ""],
    ["管理端", "系统状态概览", "运行中任务、失败统计等健康信息", "已完成", "100%", "GET /system/status"],
    ["管理端", "操作审计日志", "参数修改、回测创建等操作留痕", "已完成", "100%", "operation_logs · urt_*"],
    ["管理端", "管理端选股结果", "与网站端同源的选股结果查看", "未实现", "0%", "配置页有试算，无独立结果中心"],
    # ── 十、推送与报告 ──
    ["推送报告", "URT 每日信号推送", "定时推送自选 A 股 URT 信号日报", "已完成", "95%", "urt_daily；港股不进日报"],
    ["推送报告", "推送任务配置", "管理端配置 URT 日报类型与收件人", "已完成", "95%", ""],
    # ── 十一、部署运维 ──
    ["部署运维", "预计算调度配置", "环境变量开关与 cron 时间可配", "已完成", "95%", ""],
    ["部署运维", "数据库迁移完整性", "主业务表具备独立迁移脚本", "已完成", "100%", "add_urt_core_tables / trade_tables 等"],
]

flow_rows = [
    ["业务环节", "能力要点", "完成状态", "说明"],
    ["① 数据准备", "日 K 采集 → 候选池过滤 → 现算 MA/量能/连阳", "已完成", "A/港行情表；不依赖独立共振指标表"],
    ["② 策略计算", "硬筛 → 百分制得分 → 买点判定 → 明细拆解", "已完成", "risk_exit 已接入回测"],
    ["③ 选股筛选", "多范围 → 过滤排序 → 展示导出", "已完成", "含港股；缺 ETF/版本观察股池"],
    ["④ 信号缓存", "定时 A/港预计算 → 缓存表 → 选股优先读库", "已完成", ""],
    ["⑤ 历史追溯", "区间查询 → 版本切换 → 异步重算 → 明细页 → 单股回测", "已完成", ""],
    ["⑥ 批量回测", "任务创建 → 异步执行 → 报告/CSV/xlsx", "已完成", "hit_rate 默认；risk_exit 可选；仅 A 股"],
    ["⑦ 交易执行", "选股信号 → 交易观察 → 正式交易", "已完成", "非 GMS 式版本观察股池"],
    ["⑧ 运营推送", "urt_daily → 定时推送 → 休市跳过", "已完成", "仅自选 A 股"],
]

todo_rows = [
    ["类别", "优化项", "当前状态", "优先级", "说明"],
    ["体验", "选股页参数版本下拉", "已完成", "—", "前台 strategy-configs + config_id"],
    ["回测", "接入止损/连跌/回撤离场模式", "已完成", "—", "exit_mode=risk_exit"],
    ["交易", "交易观察股与正式交易台账", "已完成", "—", "独立 URT 表与路由"],
    ["运营", "收盘微信/邮件 URT 日报", "已完成", "—", "urt_daily；仅自选 A 股"],
    ["市场", "港股选股与预计算", "已完成", "—", "scope=hk + 约 17:20 预计算"],
    ["市场", "ETF 独立选股池", "计划中", "中", "待产品确认"],
    ["市场", "GMS 式版本观察股池", "计划中", "中", "与现有交易观察关系待确认"],
    ["市场", "港股单股/日报/回测扩展", "待确认", "中", "见对比文 §4.2 H"],
    ["运维", "Admin 审计日志与系统状态", "已完成", "低", "Phase 5 已落地"],
    ["体验", "个股详情页 URT 卡片", "已完成", "低", "stock.js 调 urt-score-detail"],
    ["工程", "主表独立 migrations 补齐", "已完成", "低", "add_urt_core_tables.py"],
]

doc_rows = [
    ["文档名称", "内容概要"],
    ["URT 策略实现设计", "策略定义、得分规则、API、权限、二期预计算与回测"],
    ["URT 策略交易回测说明", "命中率回测买卖与观察期规则"],
    ["URT 与 GMS 功能对比及技术方案", "模块对照、已落地/待办技术方案（主交付对照文）"],
    ["URT 上升趋势业务简化版", "业务口径（以代码为准）"],
    ["URT 上升趋势策略落地计划", "一期/二期范围与后续项（可能滞后）"],
    ["GMS 策略功能模块完成列表", "对照基准（Excel）"],
    ["GMS 策略实现设计", "GMS 架构参考"],
    ["GMS 回测管理中心手册", "管理端回测交互参考"],
]


def style_sheet(ws, freeze=True):
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Microsoft YaHei", size=10, color=THEME["text"])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=THEME["header"])
        cell.font = Font(name="Microsoft YaHei", size=10, color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.auto_filter.ref = ws.dimensions


def status_fill(status):
    s = str(status)
    if "已完成" in s:
        return THEME["done"]
    if "部分" in s:
        return THEME["partial"]
    if "进行" in s or "待验证" in s:
        return THEME["doing"]
    if "计划" in s:
        return THEME["planned"]
    if "未实现" in s:
        return THEME["todo"]
    return "FFFFFF"


def write_table(wb, title, rows, widths=None, status_col=None):
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    if status_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col)
            cell.fill = PatternFill("solid", fgColor=status_fill(cell.value))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    else:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 28 if row == 1 else 40
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)

    write_table(wb, "概览", [["项目", "内容", "说明"], *overview_rows], widths=[18, 42, 72])

    write_table(
        wb,
        "功能模块完成列表",
        module_rows,
        widths=[16, 22, 52, 14, 12, 32],
        status_col=4,
    )

    write_table(
        wb,
        "业务链路完成情况",
        flow_rows,
        widths=[18, 48, 14, 48],
        status_col=3,
    )

    write_table(
        wb,
        "待优化事项",
        todo_rows,
        widths=[14, 34, 14, 12, 52],
        status_col=3,
    )

    write_table(wb, "配套文档", doc_rows, widths=[36, 62])

    ws = wb["概览"]
    start = ws.max_row + 3
    data = module_rows[1:]
    summary = [
        ["统计项", "数量"],
        ["功能项合计", len(data)],
        ["已完成", sum(1 for r in data if r[3] == "已完成")],
        ["部分完成", sum(1 for r in data if r[3] == "部分完成")],
        ["未实现", sum(1 for r in data if r[3] == "未实现")],
        ["一级模块数", len({r[0] for r in data})],
    ]
    for row in summary:
        ws.append(row)
    for row in range(start, ws.max_row + 1):
        for col in range(1, 3):
            c = ws.cell(row, col)
            c.border = Border(
                left=Side(style="thin", color="D0D7DE"),
                right=Side(style="thin", color="D0D7DE"),
                top=Side(style="thin", color="D0D7DE"),
                bottom=Side(style="thin", color="D0D7DE"),
            )
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.font = Font(name="Microsoft YaHei", size=10, bold=(row == start))
            if row == start:
                c.fill = PatternFill("solid", fgColor=THEME["subheader"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
        print(OUT)
    except PermissionError:
        alt = OUT.with_name(f"URT上升趋势策略功能模块列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(alt)
        print(f"原文件被占用，已另存为: {alt}")


if __name__ == "__main__":
    build()
